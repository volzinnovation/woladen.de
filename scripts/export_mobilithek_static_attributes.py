#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import requests
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.build_data import fetch_mobilithek_access_token
from scripts.build_mobilithek_afir_configs import (
    SEARCH_PAGE_SIZE,
    classify_feed_kind,
    content_data_entry,
    fetch_offer_metadata,
    fetch_static_payload_with_probe,
    is_charging_related_offer,
    is_test_offer,
    offer_access_mode,
    probe_publication_file_access,
    search_mobilithek_offers,
)

OUTPUT_DIR = REPO_ROOT / "output" / "spreadsheet"

SPECIAL_SEGMENT_LABELS = {
    "payload": "Payload",
    "aegiEnergyInfrastructureTablePublication": "AFIR energy infrastructure table publication",
    "energyInfrastructureTable[]": "Energy infrastructure table",
    "energyInfrastructureSite[]": "Charging site",
    "energyInfrastructureStation[]": "Charging station",
    "refillPoint[]": "Charge point",
    "connector[]": "Connector",
    "electricEnergy": "Electric energy details",
    "deliveryUnit": "Delivery unit",
    "pointByCoordinates": "Coordinates block",
    "pointCoordinates": "Point coordinates",
    "locationReference": "Location reference",
    "operatingHours": "Operating hours",
    "operator": "Operator",
    "helpdesk": "Helpdesk",
    "supplementalFacility": "Supplemental facility",
    "additionalInformation": "Additional information",
    "authenticationAndIdentificationMethods": "Authentication and identification methods",
    "userInterfaceLanguage[]": "User interface language",
    "name": "Name",
    "values[]": "Localized value",
    "value": "Value",
    "idG": "Global DATEX identifier",
    "versionG": "DATEX version number",
    "numberOfRefillPoints": "Number of charge points",
    "totalMaximumPower": "Total maximum power",
    "currentType": "Current type",
    "availableChargingPower": "Available charging power",
    "availableVoltage": "Available voltage",
    "numberOfConnectors": "Number of connectors",
    "externalIdentifier": "External identifier",
    "countryCode": "Country code",
    "country": "Country",
    "street": "Street",
    "houseNumber": "House number",
    "postcode": "Postcode",
    "city": "City",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "isGreenEnergy": "Green energy flag",
    "pricePerKWh": "Price per kWh",
    "pricePerMinute": "Price per minute",
    "paymentMethod[]": "Payment method",
    "paymentMethod": "Payment method",
    "currency": "Currency",
    "lang": "Language code",
    "text": "Text",
    "taxIncludedInPrice": "Tax included in price",
    "taxRate": "Tax rate",
    "phoneNumber": "Phone number",
    "serviceType": "Service type",
    "accessibility": "Accessibility",
    "openingTimes": "Opening times",
    "openingTimeSpecifications": "Opening time specification",
    "timePeriodByHour": "Time period by hour",
    "periodName": "Period name",
    "lastUpdated": "Last updated timestamp",
    "coordinatesForDisplay": "Display coordinates",
    "locPointLocation": "Point location",
    "locLocationExtensionG": "Location extension",
    "FacilityLocation": "Facility location",
    "addressLine[]": "Address line",
}

ENTITY_LEVEL_PATTERNS = (
    ("connector", "connector[]"),
    ("charge_point", "refillPoint[]"),
    ("station", "energyInfrastructureStation[]"),
    ("site", "energyInfrastructureSite[]"),
)


@dataclass
class AttributeStat:
    provider: str
    provider_uid: str
    publisher: str
    publication_id: str
    publication_title: str
    attribute_path: str
    description: str
    entity_level: str
    occurrences: int = 0
    non_null_occurrences: int = 0
    value_types: set[str] = field(default_factory=set)
    sample_values: list[str] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export a flat static AFIR attribute inventory per provider from reachable Mobilithek subscriptions."
    )
    parser.add_argument("--search-term", default="AFIR")
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=OUTPUT_DIR / "afir_static_attributes.csv",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=OUTPUT_DIR / "afir_static_attributes.xlsx",
    )
    parser.add_argument(
        "--output-provider-csv",
        type=Path,
        default=OUTPUT_DIR / "afir_static_provider_summary.csv",
    )
    return parser.parse_args()


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def prettify_segment(segment: str) -> str:
    if segment in SPECIAL_SEGMENT_LABELS:
        return SPECIAL_SEGMENT_LABELS[segment]

    has_array = segment.endswith("[]")
    base = segment[:-2] if has_array else segment
    spaced = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", base)
    spaced = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", spaced)
    spaced = spaced.replace("_", " ").replace("-", " ")
    label = " ".join(spaced.split()).strip()
    if not label:
        label = base
    if label.lower() == "id g":
        label = "Global DATEX identifier"
    elif label.lower() == "version g":
        label = "DATEX version number"
    elif label.lower().startswith("aegi "):
        label = "AFIR " + label[5:]
    return f"{label} list" if has_array else label[:1].upper() + label[1:]


def describe_attribute_path(path: str) -> str:
    segments = [segment for segment in path.split(".") if segment]
    labels = [prettify_segment(segment) for segment in segments if segment != "payload"]
    if not labels:
        return "Payload attribute"
    if len(labels) == 1:
        return labels[0]
    leaf = labels[-1]
    parent = labels[-2]
    context = labels[:-1]
    if leaf == "Value":
        if parent == "Localized value" and len(labels) >= 3:
            target_label = labels[-3]
            context_labels = labels[:-3]
            if target_label == "Text" and len(labels) >= 4:
                target_label = labels[-4]
                context_labels = labels[:-4]
            return (
                f"Localized text value for {target_label} in {' > '.join(context_labels)}"
                if context_labels
                else f"Localized text value for {target_label}"
            )
        return f"{parent} value in {' > '.join(labels[:-2])}" if len(labels) > 2 else f"{parent} value"
    if leaf == "Language code":
        if parent == "Localized value" and len(labels) >= 3:
            target_label = labels[-3]
            context_labels = labels[:-3]
            if target_label == "Text" and len(labels) >= 4:
                target_label = labels[-4]
                context_labels = labels[:-4]
            return (
                f"Language code for localized {target_label} text in {' > '.join(context_labels)}"
                if context_labels
                else f"Language code for localized {target_label} text"
            )
        return f"Language code for {parent} in {' > '.join(labels[:-2])}" if len(labels) > 2 else f"Language code for {parent}"
    if leaf == "Order":
        return f"Ordering index for {parent} in {' > '.join(labels[:-2])}" if len(labels) > 2 else f"Ordering index for {parent}"
    return f"{leaf} in {' > '.join(context)}"


def infer_entity_level(path: str) -> str:
    for level, marker in ENTITY_LEVEL_PATTERNS:
        if marker in path:
            return level
    return "publication"


def classify_value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__


def sample_value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)
    text = " ".join(text.split())
    return text[:200]


def walk_leaf_attributes(value: Any, path: str = "") -> list[tuple[str, Any]]:
    leaves: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key in sorted(value.keys()):
            child_path = f"{path}.{key}" if path else key
            leaves.extend(walk_leaf_attributes(value[key], child_path))
        return leaves
    if isinstance(value, list):
        list_path = f"{path}[]" if path else "[]"
        if not value:
            leaves.append((list_path, None))
            return leaves
        for item in value:
            leaves.extend(walk_leaf_attributes(item, list_path))
        return leaves
    leaves.append((path, value))
    return leaves


def update_attribute_stats(
    stats: dict[tuple[str, str], AttributeStat],
    *,
    provider: str,
    provider_uid: str,
    publisher: str,
    publication_id: str,
    publication_title: str,
    payload: dict[str, Any],
) -> None:
    for path, value in walk_leaf_attributes(payload):
        key = (publication_id, path)
        attribute = stats.get(key)
        if attribute is None:
            attribute = AttributeStat(
                provider=provider,
                provider_uid=provider_uid,
                publisher=publisher,
                publication_id=publication_id,
                publication_title=publication_title,
                attribute_path=path,
                description=describe_attribute_path(path),
                entity_level=infer_entity_level(path),
            )
            stats[key] = attribute

        attribute.occurrences += 1
        value_type = classify_value_type(value)
        attribute.value_types.add(value_type)
        if value is not None and value != "":
            attribute.non_null_occurrences += 1
            sample = sample_value_text(value)
            if sample and sample not in attribute.sample_values and len(attribute.sample_values) < 3:
                attribute.sample_values.append(sample)


def fetch_static_offers(session: requests.Session, *, access_token: str, search_term: str) -> list[dict[str, Any]]:
    offers_page = search_mobilithek_offers(
        session,
        search_term=search_term,
        page=0,
        size=SEARCH_PAGE_SIZE,
        access_token=access_token,
    )
    all_offers = offers_page.get("content") or []
    static_offers: list[dict[str, Any]] = []

    for offer in all_offers:
        publication_id = str(offer.get("publicationId") or "").strip()
        if not publication_id:
            continue

        metadata = fetch_offer_metadata(session, publication_id, access_token=access_token)
        content = content_data_entry(metadata)
        title = str(metadata.get("title") or offer.get("title") or "")
        if is_test_offer(metadata, fallback_title=title):
            continue
        if not is_charging_related_offer(metadata, search_offer=offer):
            continue
        publisher = str((((metadata.get("agents") or {}).get("publisher") or {}).get("name")) or "")
        feed_kind = classify_feed_kind(metadata, fallback_title=title)
        if feed_kind != "static":
            continue

        access_probe = probe_publication_file_access(
            session,
            access_token=access_token,
            publication_id=publication_id,
        )
        static_offers.append(
            {
                "provider": publisher or title,
                "provider_uid": re.sub(r"[^a-z0-9]+", "_", (publisher or title).lower()).strip("_"),
                "publisher": publisher,
                "publication_id": publication_id,
                "title": title,
                "access_mode": offer_access_mode(metadata),
                "file_access_probe": access_probe,
                "media_type": str(content.get("mediaType") or ""),
                "data_model": str(content.get("dataModel") or ""),
                "schema_profile_name": str(content.get("schemaProfileName") or ""),
                "delta_delivery": bool(content.get("deltaDelivery")),
            }
        )
    return static_offers


def build_attribute_rows(attribute_stats: dict[tuple[str, str], AttributeStat]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for attribute in sorted(
        attribute_stats.values(),
        key=lambda item: (item.provider.lower(), item.attribute_path.lower()),
    ):
        rows.append(
            {
                "provider": attribute.provider,
                "attribute_path": attribute.attribute_path,
                "description": attribute.description,
                "provider_uid": attribute.provider_uid,
                "publisher": attribute.publisher,
                "publication_id": attribute.publication_id,
                "publication_title": attribute.publication_title,
                "entity_level": attribute.entity_level,
                "value_types": ", ".join(sorted(attribute.value_types)),
                "occurrences": attribute.occurrences,
                "non_null_occurrences": attribute.non_null_occurrences,
                "sample_values": " | ".join(attribute.sample_values),
            }
        )
    return rows


def autosize_and_style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_index = column_cells[0].column
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)

    workbook.save(path)


def write_outputs(
    *,
    attribute_rows: list[dict[str, Any]],
    provider_rows: list[dict[str, Any]],
    output_csv: Path,
    output_xlsx: Path,
    output_provider_csv: Path,
) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    output_provider_csv.parent.mkdir(parents=True, exist_ok=True)

    attributes_df = pd.DataFrame(attribute_rows)
    provider_df = pd.DataFrame(provider_rows)

    attributes_df.to_csv(output_csv, index=False)
    provider_df.to_csv(output_provider_csv, index=False)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        attributes_df.to_excel(writer, sheet_name="attributes", index=False)
        provider_df.to_excel(writer, sheet_name="providers", index=False)

    autosize_and_style_workbook(output_xlsx)


def main() -> None:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    args = parse_args()

    session = requests.Session()
    access_token = fetch_mobilithek_access_token(session)
    if not access_token:
        raise RuntimeError("missing Mobilithek credentials; set MOBILITHEK_USERNAME and MOBILITHEK_PASSWORD")

    static_offers = fetch_static_offers(session, access_token=access_token, search_term=args.search_term)
    static_offers = [
        offer for offer in static_offers if bool((offer.get("file_access_probe") or {}).get("is_accessible"))
    ]
    attribute_stats: dict[tuple[str, str], AttributeStat] = {}
    provider_rows: list[dict[str, Any]] = []
    attribute_count_by_publication: defaultdict[str, int] = defaultdict(int)

    for offer in static_offers:
        access_probe = offer["file_access_probe"]
        is_accessible = True
        fetch_status = "not_accessible"
        fetched = False
        leaf_count = 0

        payload, resolved_access_mode, fetch_error = fetch_static_payload_with_probe(
            session,
            publication_id=offer["publication_id"],
            preferred_access_mode=offer["access_mode"],
            access_token=access_token,
        )
        offer["access_mode"] = resolved_access_mode
        if payload is None:
            fetch_status = fetch_error or "fetch_failed"
        else:
            fetched = True
            fetch_status = "ok"
            leaf_count = len(walk_leaf_attributes(payload))
            update_attribute_stats(
                attribute_stats,
                provider=offer["provider"],
                provider_uid=offer["provider_uid"],
                publisher=offer["publisher"],
                publication_id=offer["publication_id"],
                publication_title=offer["title"],
                payload=payload,
            )

        provider_rows.append(
            {
                "checked_at": utc_now_iso(),
                "provider": offer["provider"],
                "provider_uid": offer["provider_uid"],
                "publisher": offer["publisher"],
                "publication_id": offer["publication_id"],
                "publication_title": offer["title"],
                "is_accessible": is_accessible,
                "fetch_status": fetch_status,
                "fetched": fetched,
                "attribute_row_count": 0,
                "leaf_value_count": leaf_count,
                "access_mode_used": offer["access_mode"],
                "file_access_status_code": access_probe.get("status_code"),
                "data_model": offer["data_model"],
                "schema_profile_name": offer["schema_profile_name"],
                "media_type": offer["media_type"],
                "delta_delivery": offer["delta_delivery"],
            }
        )

    attribute_rows = build_attribute_rows(attribute_stats)
    for row in attribute_rows:
        attribute_count_by_publication[row["publication_id"]] += 1
    for row in provider_rows:
        row["attribute_row_count"] = attribute_count_by_publication.get(row["publication_id"], 0)

    write_outputs(
        attribute_rows=attribute_rows,
        provider_rows=provider_rows,
        output_csv=args.output_csv,
        output_xlsx=args.output_xlsx,
        output_provider_csv=args.output_provider_csv,
    )

    print(
        json.dumps(
            {
                "generated_at": utc_now_iso(),
                "provider_count": len(provider_rows),
                "accessible_provider_count": len(provider_rows),
                "fetched_provider_count": sum(1 for row in provider_rows if row["fetched"]),
                "attribute_row_count": len(attribute_rows),
                "output_csv": str(args.output_csv),
                "output_provider_csv": str(args.output_provider_csv),
                "output_xlsx": str(args.output_xlsx),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
