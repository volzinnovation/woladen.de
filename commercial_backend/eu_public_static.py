from __future__ import annotations

import gzip
import io
import json
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from typing import Any, Iterable

from .stream_utils import text_stream_from_binary_stream as _text_stream_util

CY_TRAFFIC4CYPRUS_SOURCE_UID = "cy_traffic4cyprus_seed"
CY_TRAFFIC4CYPRUS_URL = "https://fixcyprus.cy/gnosis/open/api/nap/datasets/electric_vehicle_chargers/"

CZ_MPO_REGISTER_SOURCE_UID = "cz_dopravniinfo_seed"
CZ_MPO_REGISTER_URL = (
    "https://mpo.gov.cz/assets/cz/energetika/statistika/"
    "statistika-a-evidence-cerpacich-a-dobijecich-stanic/2026/4/"
    "Seznam_verDS_2026_03_31_v00.xlsx"
)

GR_ELECTROKINISI_SOURCE_UID = "gr_nap_electrokinisi_seed"
GR_IDRO_DYNAMIC_SOURCE_UID = "gr_idro_dynamic_json"
GR_IDRO_PUBLIC_DATA_PAGE_URL = "https://electrokinisi.yme.gov.gr/public/HelpMyfah/PublicData/"
GR_IDRO_STATIC_ZIP_URL = (
    "https://electrokinisi.yme.gov.gr/public/static_files/GR.IDRO.static.data.latest.json.zip"
)
GR_IDRO_DYNAMIC_ZIP_URL = (
    "https://electrokinisi.yme.gov.gr/public/static_files/GR.IDRO.dynamic.data.latest.json.zip"
)
GR_ELECTROKINISI_PAGE_URL = "https://electrokinisi.yme.gov.gr/public/ChargingPoints/"
GR_ELECTROKINISI_LOCATIONS_URL = "https://electrokinisi.yme.gov.gr/myfah-api/openApi/GetPLocations"
GR_ELECTROKINISI_LOCATION_URL = "https://electrokinisi.yme.gov.gr/myfah-api/openApi/GetLocation"

LT_LEGACY_EV_LOCATIONS_SOURCE_UID = "lt_eismoinfo_seed"
LT_LEGACY_EV_LOCATIONS_URL = "https://ev.lakd.lt/lt/api/locations/all"
LT_VIA_LIETUVA_DATEX_TABLE_SOURCE_UID = "lt_vialietuva_datex_static"
LT_VIA_LIETUVA_DATEX_STATUS_SOURCE_UID = "lt_vialietuva_datex_status"
LT_VIA_LIETUVA_OCPI_STATIC_SOURCE_UID = "lt_vialietuva_ocpi_locations_static"
LT_VIA_LIETUVA_OCPI_DYNAMIC_SOURCE_UID = "lt_vialietuva_ocpi_locations_dynamic"
LT_VIA_LIETUVA_DATEX_TABLE_URL = "https://ev.vialietuva.lt/publicdata/EnergyInfrastructureTablePublication"
LT_VIA_LIETUVA_DATEX_STATUS_URL = "https://ev.vialietuva.lt/publicdata/EnergyInfrastructureStatusPublication"
LT_VIA_LIETUVA_OCPI_LOCATIONS_URL = "https://ev.vialietuva.lt/ocpi/2.3.0/locations"
LT_EV_LOCATIONS_SOURCE_UID = LT_VIA_LIETUVA_DATEX_TABLE_SOURCE_UID
LT_EV_STATUS_SOURCE_UID = LT_VIA_LIETUVA_DATEX_STATUS_SOURCE_UID
LT_EV_LOCATIONS_URL = LT_VIA_LIETUVA_DATEX_TABLE_URL
LT_EV_STATUS_URL = LT_VIA_LIETUVA_DATEX_STATUS_URL

LU_CHARGING_STATIONS_SOURCE_UID = "lu_data_public_seed"
LU_CHARGING_STATIONS_WFS_URL = (
    "https://wms.inspire.geoportail.lu/geoserver/tn/wfs?"
    "service=WFS&version=2.0.0&request=GetFeature&"
    "typeNames=tn:TN.RoadTransportNetwork.RoadNode.ElectricalChargingStations&"
    "outputFormat=json&srsName=EPSG:4326&count=10000"
)

MT_CHARGING_POINTS_SOURCE_UID = "mt_geoservices_seed"
MT_CHARGING_POINTS_URL = (
    "https://geoservices.transport.gov.mt/arcgis/rest/services/"
    "TM_Maps/eGIS_TransportFeatures/MapServer/4/query?"
    "f=geojson&where=1%3D1&outFields=%2A&returnGeometry=true&outSR=4326"
)

ES_DGT_ELECTROLINERAS_SOURCE_UID = "es_dgt_electrolineras_datex_static"
ES_DGT_ELECTROLINERAS_URL = (
    "https://infocar.dgt.es/datex2/v3/miterd/"
    "EnergyInfrastructureTablePublication/electrolineras.xml"
)

NOBIL_DATADUMP_URL = "https://nobil.no/api/server/datadump.php"
NOBIL_REALTIME_ENDPOINT_URL = "https://api.data.enova.no/nobil/real-time/v1/Realtime"
NOBIL_REALTIME_TOKEN_METHOD = "POST"
NO_NOBIL_STATIC_SOURCE_UID = "no_nobil_static_datadump"
SE_NOBIL_STATIC_SOURCE_UID = "se_nobil_static_datadump"
NO_NOBIL_REALTIME_SOURCE_UID = "no_nobil_realtime"
SE_NOBIL_REALTIME_SOURCE_UID = "se_nobil_realtime"


@dataclass(frozen=True)
class PublicStaticSourceSpec:
    key: str
    country_code: str
    source_uid: str
    display_name: str
    url: str
    suffix: str
    accept: str
    content_type: str
    source_kind: str = "open_static_no_auth_source"
    task_kind: str = "parse_static_payload"


PUBLIC_STATIC_SOURCES: dict[str, PublicStaticSourceSpec] = {
    "cy": PublicStaticSourceSpec(
        key="cy",
        country_code="CY",
        source_uid=CY_TRAFFIC4CYPRUS_SOURCE_UID,
        display_name="CY Traffic4Cyprus/FixCyprus electric vehicle chargers DATEX II",
        url=CY_TRAFFIC4CYPRUS_URL,
        suffix=".xml",
        accept="application/xml,text/xml,*/*",
        content_type="application/xml",
    ),
    "cz": PublicStaticSourceSpec(
        key="cz",
        country_code="CZ",
        source_uid=CZ_MPO_REGISTER_SOURCE_UID,
        display_name="CZ MPO public charging-station register XLSX",
        url=CZ_MPO_REGISTER_URL,
        suffix=".xlsx",
        accept="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
    "gr": PublicStaticSourceSpec(
        key="gr",
        country_code="GR",
        source_uid=GR_ELECTROKINISI_SOURCE_UID,
        display_name="GR Electrokinisi IDRO static charging-station JSON ZIP",
        url=GR_IDRO_STATIC_ZIP_URL,
        suffix=".json.zip",
        accept="application/zip,application/x-zip-compressed,application/octet-stream,*/*",
        content_type="application/zip",
    ),
    "gr-dynamic": PublicStaticSourceSpec(
        key="gr-dynamic",
        country_code="GR",
        source_uid=GR_IDRO_DYNAMIC_SOURCE_UID,
        display_name="GR Electrokinisi IDRO dynamic charging-status JSON ZIP",
        url=GR_IDRO_DYNAMIC_ZIP_URL,
        suffix=".json.zip",
        accept="application/zip,application/x-zip-compressed,application/octet-stream,*/*",
        content_type="application/zip",
        source_kind="open_dynamic_no_auth_source",
        task_kind="parse_dynamic_payload",
    ),
    "lt": PublicStaticSourceSpec(
        key="lt",
        country_code="LT",
        source_uid=LT_EV_LOCATIONS_SOURCE_UID,
        display_name="LT Via Lietuva DATEX II public charging infrastructure table",
        url=LT_EV_LOCATIONS_URL,
        suffix=".xml",
        accept="application/xml,text/xml,*/*",
        content_type="application/xml",
    ),
    "lt-dynamic": PublicStaticSourceSpec(
        key="lt-dynamic",
        country_code="LT",
        source_uid=LT_EV_STATUS_SOURCE_UID,
        display_name="LT Via Lietuva DATEX II public charging infrastructure statuses",
        url=LT_EV_STATUS_URL,
        suffix=".xml",
        accept="application/xml,text/xml,*/*",
        content_type="application/xml",
        source_kind="open_dynamic_no_auth_source",
        task_kind="parse_dynamic_payload",
    ),
    "lu": PublicStaticSourceSpec(
        key="lu",
        country_code="LU",
        source_uid=LU_CHARGING_STATIONS_SOURCE_UID,
        display_name="LU public electrical charging stations WFS GeoJSON",
        url=LU_CHARGING_STATIONS_WFS_URL,
        suffix=".geojson",
        accept="application/json,*/*",
        content_type="application/json",
    ),
    "mt": PublicStaticSourceSpec(
        key="mt",
        country_code="MT",
        source_uid=MT_CHARGING_POINTS_SOURCE_UID,
        display_name="MT Transport Malta eGIS Charging Points ArcGIS GeoJSON",
        url=MT_CHARGING_POINTS_URL,
        suffix=".geojson",
        accept="application/geo+json,application/json,*/*",
        content_type="application/geo+json",
    ),
    "es": PublicStaticSourceSpec(
        key="es",
        country_code="ES",
        source_uid=ES_DGT_ELECTROLINERAS_SOURCE_UID,
        display_name="ES DGT electrolineras DATEX II static charging infrastructure XML",
        url=ES_DGT_ELECTROLINERAS_URL,
        suffix=".xml",
        accept="application/xml,text/xml,*/*",
        content_type="application/xml",
    ),
    "no-nobil": PublicStaticSourceSpec(
        key="no-nobil",
        country_code="NO",
        source_uid=NO_NOBIL_STATIC_SOURCE_UID,
        display_name="NO NOBIL API v3 static charging-station datadump",
        url=NOBIL_DATADUMP_URL,
        suffix=".json",
        accept="application/json,*/*",
        content_type="application/json",
        source_kind="credentialed_static_source",
    ),
    "se-nobil": PublicStaticSourceSpec(
        key="se-nobil",
        country_code="SE",
        source_uid=SE_NOBIL_STATIC_SOURCE_UID,
        display_name="SE NOBIL API v3 static charging-station datadump",
        url=NOBIL_DATADUMP_URL,
        suffix=".json",
        accept="application/json,*/*",
        content_type="application/json",
        source_kind="credentialed_static_source",
    ),
}


def _text(value: Any) -> str:
    return str(value or "").strip()


def _safe_id(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode("ascii").lower()
    cleaned = "".join(ch if ch.isalnum() or ch in {"_", "-", "*"} else "-" for ch in text)
    while "--" in cleaned:
        cleaned = cleaned.replace("--", "-")
    return cleaned.strip("-") or "unknown"


def _float_or_none(value: Any) -> float | None:
    text = _text(value).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _price_scalar(value: float) -> str:
    return f"{float(value):.6f}".rstrip("0").rstrip(".")


def _euro_amount(value: float) -> str:
    return f"{round(float(value) + 0.000000001, 2):.2f}".replace(".", ",")


def _lt_price_fields(price_text: Any) -> dict[str, Any]:
    source_text = _text(price_text)
    if not source_text:
        return {}
    normalized = source_text.casefold()
    if normalized in {"nemokama", "free", "gratis", "0", "0 eur", "0 €"}:
        return {
            "price_display": "nemokama",
            "price_currency": "EUR",
            "price_energy_eur_kwh_min": "0",
            "price_energy_eur_kwh_max": "0",
            "price_time_eur_min_min": None,
            "price_time_eur_min_max": None,
            "price_quality": "source_lt_connector_price_free",
            "price_complex": False,
            "price_source_text": source_text,
        }
    energy_values: list[float] = []
    minute_values: list[float] = []
    fixed_values: list[float] = []
    for match in re.finditer(r"(?P<value>\d+(?:[.,]\d+)?)\s*(?:€|eur)?\s*/?\s*kwh", source_text, flags=re.IGNORECASE):
        value = _float_or_none(match.group("value"))
        if value is not None:
            energy_values.append(value)
    for match in re.finditer(r"(?P<value>\d+(?:[.,]\d+)?)\s*(?:€|eur)?\s*/?\s*(?:min|minute)", source_text, flags=re.IGNORECASE):
        value = _float_or_none(match.group("value"))
        if value is not None:
            minute_values.append(value)
    for match in re.finditer(r"(?<!/)\b(?P<value>\d+(?:[.,]\d+)?)\s*(?:€|eur)\b", source_text, flags=re.IGNORECASE):
        value = _float_or_none(match.group("value"))
        if value is not None:
            fixed_values.append(value)
    if not energy_values and not minute_values:
        return {"price_source_text": source_text}
    energy_min = min(energy_values) if energy_values else None
    energy_max = max(energy_values) if energy_values else None
    minute_min = min(minute_values) if minute_values else None
    minute_max = max(minute_values) if minute_values else None
    complex_tariff = bool(minute_values or fixed_values or "+" in source_text)
    display = ""
    if energy_min is not None:
        display = f"ab {_euro_amount(energy_min)} €/kWh" if complex_tariff else f"{_euro_amount(energy_min)} €/kWh"
    elif minute_min is not None:
        display = f"ab {_euro_amount(minute_min)} €/min" if complex_tariff else f"{_euro_amount(minute_min)} €/min"
    return {
        "price_display": display,
        "price_currency": "EUR",
        "price_energy_eur_kwh_min": _price_scalar(energy_min) if energy_min is not None else "",
        "price_energy_eur_kwh_max": _price_scalar(energy_max) if energy_max is not None else "",
        "price_time_eur_min_min": round(minute_min, 6) if minute_min is not None else None,
        "price_time_eur_min_max": round(minute_max, 6) if minute_max is not None else None,
        "price_quality": "source_lt_connector_price_complex" if complex_tariff else "source_lt_connector_price_exact",
        "price_complex": complex_tariff,
        "price_source_text": source_text,
    }


def _merge_price_field_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    priced_rows = [
        row
        for row in rows
        if row
        and (
            _text(row.get("price_display"))
            or _text(row.get("price_energy_eur_kwh_min"))
            or row.get("price_time_eur_min_min") is not None
        )
    ]
    if not priced_rows:
        source_texts = list(dict.fromkeys(_text(row.get("price_source_text")) for row in rows if _text(row.get("price_source_text"))))
        return {"price_source_text": "|".join(source_texts)} if source_texts else {}
    displays = list(dict.fromkeys(_text(row.get("price_display")) for row in priced_rows if _text(row.get("price_display"))))
    energy_values = [
        value
        for row in priced_rows
        for value in (_float_or_none(row.get("price_energy_eur_kwh_min")), _float_or_none(row.get("price_energy_eur_kwh_max")))
        if value is not None
    ]
    minute_values = [
        value
        for row in priced_rows
        for value in (_float_or_none(row.get("price_time_eur_min_min")), _float_or_none(row.get("price_time_eur_min_max")))
        if value is not None
    ]
    source_texts = list(dict.fromkeys(_text(row.get("price_source_text")) for row in rows if _text(row.get("price_source_text"))))
    currency = next((_text(row.get("price_currency")) for row in priced_rows if _text(row.get("price_currency"))), "EUR")
    complex_tariff = len(displays) > 1 or any(bool(row.get("price_complex")) for row in priced_rows)
    energy_min = min(energy_values) if energy_values else None
    energy_max = max(energy_values) if energy_values else None
    minute_min = min(minute_values) if minute_values else None
    minute_max = max(minute_values) if minute_values else None
    display = displays[0] if len(displays) == 1 else ""
    if not display and energy_min is not None and currency == "EUR":
        display = f"ab {_euro_amount(energy_min)} €/kWh" if complex_tariff else f"{_euro_amount(energy_min)} €/kWh"
    return {
        "price_display": display,
        "price_currency": currency,
        "price_energy_eur_kwh_min": _price_scalar(energy_min) if energy_min is not None and currency == "EUR" else "",
        "price_energy_eur_kwh_max": _price_scalar(energy_max) if energy_max is not None and currency == "EUR" else "",
        "price_time_eur_min_min": round(minute_min, 6) if minute_min is not None and currency == "EUR" else None,
        "price_time_eur_min_max": round(minute_max, 6) if minute_max is not None and currency == "EUR" else None,
        "price_quality": "source_price_mixed" if complex_tariff else _text(priced_rows[0].get("price_quality")),
        "price_complex": complex_tariff,
        "price_source_text": "|".join(source_texts[:5]),
    }


def _int_or_none(value: Any) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _compact_json(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)


def _lt_opening_hours(value: Any) -> str:
    if isinstance(value, dict) and value.get("twentyfourseven") is True:
        return "24/7"
    return _compact_json(value)


def _first_text(parent: ET.Element, name: str) -> str:
    for child in parent.iter():
        if child.tag.split("}", 1)[-1] == name and child.text and child.text.strip():
            return child.text.strip()
    return ""


def _child_value_text(parent: ET.Element, name: str) -> str:
    for child in parent:
        if _local_name(child) != name:
            continue
        direct = _text(child.text)
        if direct:
            return direct
        for descendant in child.iter():
            if _local_name(descendant) == "value" and _text(descendant.text):
                return _text(descendant.text)
        return ""
    return ""


def _local_name(element: ET.Element) -> str:
    return element.tag.split("}", 1)[-1]


def _text_stream_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str,
    encoding: str = "utf-8",
) -> io.TextIOWrapper:
    return _text_stream_util(raw_stream, content_encoding=content_encoding, encoding=encoding)


def _json_from_binary_stream(raw_stream: io.BufferedIOBase, *, content_encoding: str = "") -> Any:
    text_stream = _text_stream_from_binary_stream(raw_stream, content_encoding=content_encoding)
    text = text_stream.read()
    first_json = min((index for index in (text.find("{"), text.find("[")) if index >= 0), default=0)
    return json.loads(text[first_json:])


def _json_from_zip_or_binary_stream(raw_stream: io.BufferedIOBase, *, content_encoding: str = "") -> Any:
    payload = raw_stream.read()
    if content_encoding.casefold() == "gzip" or payload.startswith(b"\x1f\x8b"):
        payload = gzip.decompress(payload)
    if payload.startswith(b"PK"):
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            json_names = [name for name in archive.namelist() if name.lower().endswith(".json")]
            if not json_names:
                raise ValueError("zip_payload_missing_json_member")
            payload = archive.read(sorted(json_names)[0])
    text = payload.decode("utf-8", errors="replace")
    first_json = min((index for index in (text.find("{"), text.find("[")) if index >= 0), default=0)
    return json.loads(text[first_json:])


def _max_power_from_connectors(connectors: Iterable[dict[str, Any]]) -> float | None:
    powers: list[float] = []
    for connector in connectors:
        power = _float_or_none(connector.get("max_electric_power"))
        if power and power > 1000:
            powers.append(power / 1000.0)
            continue
        elif power:
            powers.append(power)
            continue
        voltage = _float_or_none(connector.get("max_voltage"))
        amperage = _float_or_none(connector.get("max_amperage"))
        if voltage and amperage:
            phases = 3 if "3" in _text(connector.get("power_type")) else 1
            powers.append(voltage * amperage * phases / 1000.0)
    return max(powers) if powers else None


def _direct_children(parent: ET.Element, name: str) -> list[ET.Element]:
    return [child for child in parent if _local_name(child) == name]


def _direct_child_text(parent: ET.Element, name: str) -> str:
    for child in parent:
        if _local_name(child) == name:
            return _text(child.text)
    return ""


def _values_text(parent: ET.Element) -> list[str]:
    values: list[str] = []
    for descendant in parent.iter():
        if _local_name(descendant) == "value" and _text(descendant.text):
            values.append(_text(descendant.text))
    return values


def _direct_child_value_text(parent: ET.Element, name: str) -> str:
    for child in parent:
        if _local_name(child) != name:
            continue
        direct = _text(child.text)
        if direct:
            return direct
        values = _values_text(child)
        return values[0] if values else ""
    return ""


def _es_address_parts(site: ET.Element) -> dict[str, str]:
    parts: dict[str, str] = {}
    loose_lines: list[str] = []
    for address_line in site.iter():
        if _local_name(address_line) != "addressLine":
            continue
        line = _direct_child_value_text(address_line, "text")
        if not line:
            continue
        if ":" not in line:
            loose_lines.append(line)
            continue
        label, value = line.split(":", 1)
        normalized_label = unicodedata.normalize("NFKD", label).encode("ascii", "ignore").decode("ascii").casefold()
        value = value.strip()
        if normalized_label == "direccion":
            parts["address"] = value
        elif normalized_label == "municipio":
            parts["city"] = value
        elif normalized_label == "provincia":
            parts["province"] = value
        elif normalized_label == "comunidad autonoma":
            parts["autonomous_community"] = value
    if "address" not in parts and loose_lines:
        parts["address"] = " | ".join(loose_lines)
    return parts


def _es_operator(site: ET.Element) -> tuple[str, str]:
    for child in site:
        if _local_name(child) != "operator":
            continue
        return _text(child.attrib.get("id")), _direct_child_value_text(child, "name")
    return "", ""


def _es_source_evse_id(refill_point: ET.Element, fallback: str) -> str:
    candidates = _values_text(refill_point)
    for candidate in candidates:
        if candidate.upper().startswith("ES*") or re.search(r"\bES\*[A-Z0-9]{3}\*", candidate, flags=re.IGNORECASE):
            return candidate
    for candidate in candidates:
        if candidate:
            return candidate
    return fallback


def _es_current_type(values: Iterable[str]) -> str:
    normalized = " ".join(_text(value).casefold() for value in values)
    if "dc" in normalized or "chademo" in normalized or "combo" in normalized:
        return "DC"
    if "ac" in normalized or "iec62196t2" in normalized:
        return "AC"
    return ""


def _es_connector_rows(refill_point: ET.Element) -> list[dict[str, Any]]:
    connectors: list[dict[str, Any]] = []
    for index, connector in enumerate(_direct_children(refill_point, "connector"), start=1):
        connector_type = _direct_child_text(connector, "connectorType")
        charging_mode = _direct_child_text(connector, "chargingMode")
        power = _float_or_none(_direct_child_text(connector, "maxPowerAtSocket"))
        if power is not None and power > 1000:
            power = power / 1000.0
        connector_format = _direct_child_text(connector, "connectorFormat")
        connector_id = _text(connector.attrib.get("id")) or str(index)
        connectors.append(
            {
                "connector_id": connector_id,
                "connector_type": connector_type,
                "charging_mode": charging_mode,
                "connector_format": connector_format,
                "current_type": _es_current_type((connector_type, charging_mode)),
                "max_power_kw": power,
            }
        )
    return connectors


def iter_es_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    text_stream = _text_stream_from_binary_stream(raw_stream, content_encoding=content_encoding)
    publication_time = ""
    seen_charger_ids: set[str] = set()
    for event, element in ET.iterparse(text_stream, events=("end",)):
        local_name = _local_name(element)
        if local_name == "publicationTime" and not publication_time:
            publication_time = _text(element.text)
            continue
        if local_name != "energyInfrastructureSite":
            continue

        source_station_id = _text(element.attrib.get("id"))
        if not source_station_id:
            element.clear()
            continue
        station_id = f"es:dgt:site:{_safe_id(source_station_id)}"
        station_name = _direct_child_value_text(element, "name")
        date_updated = _direct_child_text(element, "lastUpdated") or publication_time
        opening_hours = ""
        operating_hours = _direct_children(element, "operatingHours")
        if operating_hours:
            opening_hours = _text(operating_hours[0].attrib.get("id"))
            if opening_hours.casefold() in {"no disponible", "none", "unknown"}:
                opening_hours = ""
        address_parts = _es_address_parts(element)
        postal_code = _first_text(element, "postcode")
        latitude = _float_or_none(_first_text(element, "latitude"))
        longitude = _float_or_none(_first_text(element, "longitude"))
        operator_id, operator_name = _es_operator(element)
        operator_name = operator_name or operator_id

        for station in element.iter():
            if _local_name(station) != "energyInfrastructureStation":
                continue
            auth_methods = "|".join(
                dict.fromkeys(
                    _text(child.text)
                    for child in _direct_children(station, "authenticationAndIdentificationMethods")
                    if _text(child.text)
                )
            )
            station_source_id = _text(station.attrib.get("id")) or source_station_id
            for refill_point in _direct_children(station, "refillPoint"):
                refill_point_id = _text(refill_point.attrib.get("id"))
                source_evse_id = _es_source_evse_id(refill_point, refill_point_id or station_source_id)
                if not source_evse_id:
                    continue
                base_charger_id = f"es:dgt:evse:{_safe_id(source_evse_id)}"
                charger_id = base_charger_id
                duplicate_index = 1
                while charger_id in seen_charger_ids:
                    duplicate_index += 1
                    charger_id = (
                        f"{base_charger_id}:"
                        f"{_safe_id(refill_point_id or station_source_id)}:"
                        f"{duplicate_index}"
                    )
                seen_charger_ids.add(charger_id)
                connectors = _es_connector_rows(refill_point)
                connector_types = "|".join(
                    dict.fromkeys(_text(connector.get("connector_type")) for connector in connectors if _text(connector.get("connector_type")))
                )
                current_types = "|".join(
                    dict.fromkeys(_text(connector.get("current_type")) for connector in connectors if _text(connector.get("current_type")))
                )
                power_values = [
                    connector.get("max_power_kw")
                    for connector in connectors
                    if connector.get("max_power_kw") is not None
                ]
                yield {
                    "country_code": "ES",
                    "source_uid": ES_DGT_ELECTROLINERAS_SOURCE_UID,
                    "provider_uid": "es_dgt_electrolineras",
                    "station_id": station_id,
                    "charger_id": charger_id,
                    "source_station_id": source_station_id,
                    "source_evse_id": source_evse_id,
                    "connector_id": "|".join(
                        dict.fromkeys(_text(connector.get("connector_id")) for connector in connectors if _text(connector.get("connector_id")))
                    ),
                    "connector_types": connector_types,
                    "current_type": current_types,
                    "max_power_kw": max(power_values) if power_values else None,
                    "operator_name": operator_name,
                    "station_name": station_name,
                    "address": address_parts.get("address", ""),
                    "postal_code": postal_code,
                    "city": address_parts.get("city", ""),
                    "latitude": latitude,
                    "longitude": longitude,
                    "opening_hours": opening_hours,
                    "auth_methods": auth_methods,
                    "date_updated": date_updated,
                }
        element.clear()


def iter_cy_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    text_stream = _text_stream_from_binary_stream(raw_stream, content_encoding=content_encoding)
    root = ET.parse(text_stream).getroot()
    for charging_point in root.iter():
        if _local_name(charging_point) != "chargingPoint":
            continue
        source_evse_id = _first_text(charging_point, "chargingPointIdentification")
        if not source_evse_id:
            continue
        operator_name = _first_text(charging_point, "chargingPointOperatorLegalName") or _first_text(
            charging_point, "chargingPointOperator"
        )
        latitude = _float_or_none(_first_text(charging_point, "latitude"))
        longitude = _float_or_none(_first_text(charging_point, "longitude"))
        station_id = (
            f"cy:coord:{_safe_id(operator_name)}:"
            f"{'' if latitude is None else f'{latitude:.7f}'}:"
            f"{'' if longitude is None else f'{longitude:.7f}'}"
        )
        connector_types: list[str] = []
        for container in charging_point.iter():
            if _local_name(container) != "connectorTypes":
                continue
            for connector_type in container.iter():
                if _local_name(connector_type) == "connectorType" and _text(connector_type.text):
                    connector_types.append(_text(connector_type.text))
                if _local_name(connector_type) == "value" and _text(connector_type.text):
                    connector_types.append(_text(connector_type.text))
        yield {
            "country_code": "CY",
            "source_uid": CY_TRAFFIC4CYPRUS_SOURCE_UID,
            "provider_uid": "cy_traffic4cyprus",
            "station_id": station_id,
            "charger_id": f"cy:datex:evse:{_safe_id(source_evse_id)}",
            "source_station_id": station_id,
            "source_evse_id": source_evse_id,
            "operator_name": operator_name,
            "station_name": source_evse_id,
            "address": _child_value_text(charging_point, "chargingPointAddress"),
            "city": "",
            "postal_code": "",
            "latitude": latitude,
            "longitude": longitude,
            "connector_count": _int_or_none(_first_text(charging_point, "numberOfConnectors")) or 1,
            "connector_types": "|".join(dict.fromkeys(connector_types)),
            "current_type": _first_text(charging_point, "powerType").upper(),
            "max_power_kw": _float_or_none(_first_text(charging_point, "maximumPower"))
            or _float_or_none(_first_text(charging_point, "connectorPower")),
            "opening_hours": _child_value_text(charging_point, "operatingTime"),
            "date_updated": _first_text(charging_point, "creationDate"),
        }


def _cz_connector_groups(row: tuple[Any, ...]) -> list[dict[str, Any]]:
    entries: dict[int, dict[str, Any]] = {}
    for index, start in enumerate((12, 17, 22), start=1):
        current_type = _text(row[start]).upper()
        power = _float_or_none(row[start + 1])
        connector_type = _text(row[start + 2])
        connection = _text(row[start + 3])
        excludes = [int(value) for value in re.findall(r"\d+", _text(row[start + 4]))]
        if current_type or power is not None or connector_type:
            entries[index] = {
                "index": index,
                "current_type": current_type,
                "max_power_kw": power,
                "connector_type": connector_type,
                "connection": connection,
                "excludes": excludes,
            }
    if not entries:
        return []
    visited: set[int] = set()
    groups: list[list[dict[str, Any]]] = []
    for index in sorted(entries):
        if index in visited:
            continue
        stack = [index]
        group: list[dict[str, Any]] = []
        while stack:
            current = stack.pop()
            if current in visited or current not in entries:
                continue
            visited.add(current)
            group.append(entries[current])
            stack.extend(entries[current]["excludes"])
        groups.append(group)
    point_count = _int_or_none(row[11]) or len(groups)
    while point_count > 0 and len(groups) > point_count:
        groups[-2].extend(groups[-1])
        groups.pop()
    while len(groups) < point_count:
        groups.append([])
    return [
        {
            "connector_id": ";".join(str(item["index"]) for item in group) or str(index),
            "connector_type": "|".join(dict.fromkeys(item["connector_type"] for item in group if item["connector_type"])),
            "current_type": "DC" if any(item["current_type"] in {"DC", "UFC"} for item in group) else "|".join(
                dict.fromkeys(item["current_type"] for item in group if item["current_type"])
            ),
            "max_power_kw": max((item["max_power_kw"] for item in group if item["max_power_kw"] is not None), default=None),
        }
        for index, group in enumerate(groups, start=1)
    ]


def iter_cz_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    from openpyxl import load_workbook

    data = raw_stream.read()
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    for excel_row, row in enumerate(sheet.iter_rows(min_row=8, values_only=True), start=8):
        if not row or not _text(row[1]):
            continue
        address = _text(row[1])
        position = _text(row[2])
        postal_code = _text(row[3])
        city = _text(row[4])
        latitude = _float_or_none(row[6])
        longitude = _float_or_none(row[7])
        cpo_id = _text(row[8])
        operator_name = _text(row[9])
        commissioned = row[27].date().isoformat() if hasattr(row[27], "date") else _text(row[27])
        source_station_id = "|".join([f"row:{excel_row}", cpo_id, postal_code, city, address, position, commissioned])
        station_id = f"cz:mpo:{_safe_id(source_station_id)}"
        for connector_index, connector in enumerate(_cz_connector_groups(row), start=1):
            source_evse_id = f"{source_station_id}|point:{connector_index}"
            yield {
                "country_code": "CZ",
                "source_uid": CZ_MPO_REGISTER_SOURCE_UID,
                "provider_uid": "cz_mpo_register",
                "station_id": station_id,
                "charger_id": f"cz:mpo:point:{_safe_id(source_evse_id)}",
                "source_station_id": source_station_id,
                "source_evse_id": source_evse_id,
                "connector_id": connector["connector_id"],
                "connector_types": connector["connector_type"],
                "current_type": connector["current_type"],
                "max_power_kw": connector["max_power_kw"],
                "operator_name": operator_name,
                "station_name": address,
                "address": address,
                "postal_code": postal_code,
                "city": city,
                "latitude": latitude,
                "longitude": longitude,
                "opening_hours": _text(row[5]),
                "date_updated": commissioned,
            }


GR_IDRO_STATUS_MAP = {
    "AVAILABLE": "free",
    "BLOCKED": "occupied",
    "CHARGING": "occupied",
    "INOPERATIVE": "out_of_order",
    "OUTOFORDER": "out_of_order",
    "PLANNED": "unknown",
    "REMOVED": "out_of_order",
    "RESERVED": "occupied",
    "UNKNOWN": "unknown",
}


def _gr_locations_from_payload(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    locations = payload.get("Locations")
    if isinstance(locations, list):
        return [location for location in locations if isinstance(location, dict)]
    return []


def _gr_coordinates(value: Any) -> tuple[float | None, float | None]:
    if not isinstance(value, dict):
        return None, None
    return _float_or_none(value.get("latitude")), _float_or_none(value.get("longitude"))


def _gr_operator_name(loc: dict[str, Any]) -> str:
    operator = loc.get("operator") if isinstance(loc.get("operator"), dict) else {}
    suboperator = loc.get("suboperator") if isinstance(loc.get("suboperator"), dict) else {}
    owner = loc.get("owner") if isinstance(loc.get("owner"), dict) else {}
    open_api = loc.get("_openApiLocation") if isinstance(loc.get("_openApiLocation"), dict) else {}
    return (
        _text(operator.get("name"))
        or _text(suboperator.get("name"))
        or _text(open_api.get("CompanyCommercialName"))
        or _text(owner.get("name"))
        or _text(loc.get("party_id"))
    )


def _gr_opening_hours(opening_times: Any) -> str:
    if not isinstance(opening_times, dict):
        return ""
    if opening_times.get("twentyfourseven"):
        return "24/7"
    regular_hours = opening_times.get("regular_hours")
    if not isinstance(regular_hours, list):
        return ""
    spans: list[str] = []
    for item in regular_hours:
        if not isinstance(item, dict):
            continue
        weekday = _text(item.get("weekday"))
        period_begin = _text(item.get("period_begin"))
        period_end = _text(item.get("period_end"))
        if weekday and period_begin and period_end:
            spans.append(f"{weekday} {period_begin}-{period_end}")
    return "; ".join(spans)


def _gr_idro_static_rows_from_location(loc: dict[str, Any]) -> Iterable[dict[str, Any]]:
    if loc.get("publish") is False:
        return
    source_station_id = _text(loc.get("id"))
    if not source_station_id:
        return
    station_id = f"gr:electrokinisi:loc:{_safe_id(source_station_id)}"
    latitude, longitude = _gr_coordinates(loc.get("coordinates"))
    operator_name = _gr_operator_name(loc)
    opening_hours = _gr_opening_hours(loc.get("opening_times"))
    energy_mix = loc.get("energy_mix") if isinstance(loc.get("energy_mix"), dict) else {}
    green_energy = ""
    if "is_green_energy" in energy_mix:
        green_energy = "true" if energy_mix.get("is_green_energy") else "false"
    evses = [evse for evse in loc.get("evses") or [] if isinstance(evse, dict)]
    if not evses:
        yield {
            "country_code": "GR",
            "source_uid": GR_ELECTROKINISI_SOURCE_UID,
            "provider_uid": "gr_electrokinisi",
            "station_id": station_id,
            "charger_id": f"gr:electrokinisi:location-placeholder:{_safe_id(source_station_id)}",
            "source_station_id": source_station_id,
            "source_evse_id": f"{source_station_id}|location-placeholder",
            "connector_id": "",
            "connector_types": "",
            "current_type": "",
            "max_power_kw": None,
            "operator_name": operator_name,
            "station_name": _text(loc.get("name")),
            "address": _text(loc.get("address")),
            "postal_code": _text(loc.get("postal_code")),
            "city": _text(loc.get("city")),
            "latitude": latitude,
            "longitude": longitude,
            "opening_hours": opening_hours,
            "green_energy": green_energy,
            "date_updated": _text(loc.get("last_updated")),
            "public_bundle_note": "location_without_evses_in_idro_static_payload",
        }
        return
    for evse in evses:
        source_evse_id = _text(evse.get("evse_id")) or _text(evse.get("uid"))
        if not source_evse_id:
            continue
        connectors = [item for item in evse.get("connectors") or [] if isinstance(item, dict)]
        connector_types = "|".join(dict.fromkeys(_text(item.get("standard")) for item in connectors if _text(item.get("standard"))))
        current_types = "|".join(dict.fromkeys(_text(item.get("power_type")) for item in connectors if _text(item.get("power_type"))))
        connector_formats = "|".join(dict.fromkeys(_text(item.get("format")) for item in connectors if _text(item.get("format"))))
        terms = "|".join(
            dict.fromkeys(_text(item.get("terms_and_conditions")) for item in connectors if _text(item.get("terms_and_conditions")))
        )
        capabilities = "|".join(dict.fromkeys(_text(value) for value in evse.get("capabilities") or [] if _text(value)))
        row_latitude, row_longitude = _gr_coordinates(evse.get("coordinates"))
        yield {
            "country_code": "GR",
            "source_uid": GR_ELECTROKINISI_SOURCE_UID,
            "provider_uid": "gr_electrokinisi",
            "station_id": station_id,
            "charger_id": f"gr:electrokinisi:evse:{_safe_id(source_evse_id)}",
            "source_station_id": source_station_id,
            "source_evse_id": source_evse_id,
            "connector_id": "|".join(dict.fromkeys(_text(item.get("id")) for item in connectors if _text(item.get("id")))),
            "connector_types": connector_types,
            "current_type": current_types,
            "max_power_kw": _max_power_from_connectors(connectors),
            "operator_name": operator_name,
            "station_name": _text(loc.get("name")),
            "address": _text(loc.get("address")),
            "postal_code": _text(loc.get("postal_code")),
            "city": _text(loc.get("city")),
            "latitude": row_latitude if row_latitude is not None else latitude,
            "longitude": row_longitude if row_longitude is not None else longitude,
            "opening_hours": opening_hours,
            "payment_methods": "roaming" if evse.get("supports_roaming") else "",
            "auth_methods": capabilities,
            "green_energy": green_energy,
            "date_updated": _text(evse.get("last_updated")) or _text(loc.get("last_updated")),
            "connector_format": connector_formats,
            "terms_and_conditions": terms,
            "hardware_manufacturer": _text(evse.get("manufacturer")),
            "hardware_model": _text(evse.get("model_name")),
            "parking_type": _text(loc.get("parking_type")),
        }


def _iter_gr_idro_static_rows(payload: Any) -> Iterable[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for loc in _gr_locations_from_payload(payload):
        rows.extend(_gr_idro_static_rows_from_location(loc))
    rows.sort(
        key=lambda row: (
            _text(row.get("source_evse_id")).casefold(),
            _text(row.get("date_updated")),
            _text(row.get("station_id")),
        )
    )
    yield from rows


def _iter_gr_legacy_rows(payload: Any) -> Iterable[dict[str, Any]]:
    details = payload.get("details") if isinstance(payload, dict) else {}
    if not isinstance(details, dict):
        details = {}
    handled_location_ids: set[str] = set()
    for location_id, response in sorted(details.items()):
        loc = response.get("Loc") if isinstance(response, dict) else None
        if not isinstance(loc, dict):
            continue
        handled_location_ids.add(_text(location_id))
        loc = dict(loc)
        loc.setdefault("id", location_id)
        yield from _gr_idro_static_rows_from_location(loc)
    locations = payload.get("locations") if isinstance(payload, dict) else {}
    features = locations.get("features") if isinstance(locations, dict) else []
    if not isinstance(features, list):
        return
    for feature in features:
        if not isinstance(feature, dict):
            continue
        properties = feature.get("properties") if isinstance(feature.get("properties"), dict) else {}
        location_id = _text(properties.get("location_id"))
        if not location_id or location_id in handled_location_ids:
            continue
        geometry = feature.get("geometry") if isinstance(feature.get("geometry"), dict) else {}
        coords = geometry.get("coordinates") if isinstance(geometry.get("coordinates"), list) else []
        latitude = _float_or_none(coords[0]) if len(coords) >= 2 else None
        longitude = _float_or_none(coords[1]) if len(coords) >= 2 else None
        source_station_id = location_id
        station_id = f"gr:electrokinisi:loc:{_safe_id(source_station_id)}"
        yield {
            "country_code": "GR",
            "source_uid": GR_ELECTROKINISI_SOURCE_UID,
            "provider_uid": "gr_electrokinisi",
            "station_id": station_id,
            "charger_id": f"gr:electrokinisi:location-placeholder:{_safe_id(source_station_id)}",
            "source_station_id": source_station_id,
            "source_evse_id": f"{source_station_id}|location-placeholder",
            "connector_id": "",
            "connector_types": "",
            "current_type": "",
            "max_power_kw": None,
            "operator_name": _text(properties.get("Provider")),
            "station_name": _text(properties.get("LocationName")),
            "address": "",
            "postal_code": "",
            "city": "",
            "latitude": latitude,
            "longitude": longitude,
            "date_updated": _text(properties.get("LocationLastUpdatedUTC")),
            "public_bundle_note": "location_only_placeholder_due_detail_api_rate_limit",
        }


def iter_gr_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_zip_or_binary_stream(raw_stream, content_encoding=content_encoding)
    if _gr_locations_from_payload(payload):
        yield from _iter_gr_idro_static_rows(payload)
    else:
        yield from _iter_gr_legacy_rows(payload)


def iter_gr_dynamic_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_zip_or_binary_stream(raw_stream, content_encoding=content_encoding)
    rows: list[dict[str, Any]] = []
    for loc in _gr_locations_from_payload(payload):
        source_station_id = _text(loc.get("id"))
        if not source_station_id:
            continue
        station_id = f"gr:electrokinisi:loc:{_safe_id(source_station_id)}"
        location_last_updated = _text(loc.get("last_updated"))
        for evse in loc.get("evses") or []:
            if not isinstance(evse, dict):
                continue
            source_evse_id = _text(evse.get("evse_id")) or _text(evse.get("uid"))
            if not source_evse_id:
                continue
            source_status = _text(evse.get("status")).upper() or "UNKNOWN"
            connectors = [item for item in evse.get("connectors") or [] if isinstance(item, dict)]
            rows.append(
                {
                    "country_code": "GR",
                    "source_uid": GR_IDRO_DYNAMIC_SOURCE_UID,
                    "provider_uid": "gr_electrokinisi",
                    "station_id": station_id,
                    "charger_id": f"gr:electrokinisi:evse:{_safe_id(source_evse_id)}",
                    "source_station_id": source_station_id,
                    "source_evse_id": source_evse_id,
                    "source_status": source_status,
                    "availability_status": GR_IDRO_STATUS_MAP.get(source_status, "unknown"),
                    "source_observed_at": _text(evse.get("last_updated")) or location_last_updated,
                    "connector_id": "|".join(
                        dict.fromkeys(_text(item.get("id")) for item in connectors if _text(item.get("id")))
                    ),
                }
            )
    rows.sort(
        key=lambda row: (
            _text(row.get("source_evse_id")).casefold(),
            _text(row.get("source_observed_at")),
            _text(row.get("station_id")),
        )
    )
    yield from rows


NOBIL_SOURCE_UID_BY_COUNTRY = {
    "NO": NO_NOBIL_STATIC_SOURCE_UID,
    "SE": SE_NOBIL_STATIC_SOURCE_UID,
}
NOBIL_REALTIME_SOURCE_UID_BY_COUNTRY = {
    "NO": NO_NOBIL_REALTIME_SOURCE_UID,
    "SE": SE_NOBIL_REALTIME_SOURCE_UID,
}
NOBIL_PROVIDER_UID_BY_COUNTRY = {
    "NO": "no_nobil",
    "SE": "se_nobil",
}
NOBIL_COUNTRY_BY_LAND_CODE = {
    "NO": "NO",
    "NOR": "NO",
    "NORWAY": "NO",
    "N": "NO",
    "SE": "SE",
    "SWE": "SE",
    "SWEDEN": "SE",
}
NOBIL_REALTIME_STATUS_MAP = GR_IDRO_STATUS_MAP


def _nobil_station_items(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict):
        stations = payload.get("chargerstations") or payload.get("chargingstations")
        if isinstance(stations, list):
            return [station for station in stations if isinstance(station, dict)]
        if isinstance(payload.get("csmd"), dict):
            return [payload]
        return []
    if isinstance(payload, list):
        rows: list[dict[str, Any]] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            if isinstance(item.get("chargerstations"), list) or isinstance(item.get("chargingstations"), list):
                rows.extend(_nobil_station_items(item))
            elif isinstance(item.get("csmd"), dict):
                rows.append(item)
        return rows
    return []


def _nobil_attr_entries(section: Any) -> list[dict[str, Any]]:
    if isinstance(section, list):
        return [entry for entry in section if isinstance(entry, dict)]
    if not isinstance(section, dict):
        return []
    if "attrname" in section or "trans" in section:
        return [section]
    entries: list[dict[str, Any]] = []
    for key, value in sorted(section.items(), key=lambda item: str(item[0])):
        if isinstance(value, dict):
            entry = dict(value)
            entry.setdefault("_entry_key", key)
            entries.append(entry)
    return entries


def _nobil_attr_value(entry: dict[str, Any]) -> str:
    for key in ("trans", "value", "description", "name"):
        value = _text(entry.get(key))
        if value:
            return value
    value = entry.get("attrval")
    if isinstance(value, bool):
        return "true" if value else "false"
    if value not in (None, "", [], {}):
        if isinstance(value, (dict, list)):
            return _compact_json(value)
        return _text(value)
    return ""


def _nobil_attr_raw_value(entry: dict[str, Any]) -> str:
    value = entry.get("attrval")
    if isinstance(value, bool):
        return "true" if value else "false"
    if value not in (None, "", [], {}):
        if isinstance(value, (dict, list)):
            return _compact_json(value)
        return _text(value)
    return ""


def _nobil_attr_texts(
    section: Any,
    *,
    names: set[str] | None = None,
    ids: set[str] | None = None,
    prefer_attrval: bool = False,
) -> list[str]:
    wanted_names = {name.casefold() for name in names or set()}
    wanted_ids = {str(value).casefold() for value in ids or set()}
    values: list[str] = []
    for entry in _nobil_attr_entries(section):
        attr_name = _text(entry.get("attrname")).casefold()
        attr_id = (_text(entry.get("id")) or _text(entry.get("attrtypeid")) or _text(entry.get("_entry_key"))).casefold()
        if wanted_names and attr_name not in wanted_names:
            continue
        if wanted_ids and attr_id not in wanted_ids:
            continue
        value = _nobil_attr_raw_value(entry) if prefer_attrval else ""
        if not value:
            value = _nobil_attr_value(entry)
        if value:
            values.append(value)
    return list(dict.fromkeys(values))


def _nobil_first_attr_text(
    section: Any,
    *,
    names: set[str] | None = None,
    ids: set[str] | None = None,
    prefer_attrval: bool = False,
) -> str:
    values = _nobil_attr_texts(section, names=names, ids=ids, prefer_attrval=prefer_attrval)
    return values[0] if values else ""


def _nobil_connection_groups(section: Any, *, point_count: int) -> list[tuple[str, Any]]:
    if not isinstance(section, dict):
        return [(str(index), {}) for index in range(1, point_count + 1)]
    flat_entries = [
        entry
        for entry in _nobil_attr_entries(section)
        if _text(entry.get("attrname")) or _text(entry.get("attrtypeid")) or _text(entry.get("trans"))
    ]
    if flat_entries:
        if point_count > 1:
            return [(str(index), section) for index in range(1, point_count + 1)]
        return [("1", section)]
    groups: list[tuple[str, Any]] = []
    for key, value in sorted(section.items(), key=lambda item: int(item[0]) if str(item[0]).isdigit() else str(item[0])):
        if isinstance(value, dict):
            groups.append((_text(key), value))
    if groups:
        if len(groups) < point_count:
            existing_ids = {_text(group_id) for group_id, _group in groups}
            for index in range(1, point_count + 1):
                connector_id = str(index)
                if connector_id not in existing_ids:
                    groups.append((connector_id, {}))
        return groups
    return [(str(index), {}) for index in range(1, point_count + 1)]


def _nobil_opening_hours(station_attrs: Any) -> str:
    values = _nobil_attr_texts(station_attrs, names={"Open 24h"}, ids={"24"})
    for value in values:
        if value.casefold() in {"yes", "true", "1", "ja"}:
            return "24/7"
    return ""


def _nobil_coordinates(value: Any) -> tuple[float | None, float | None]:
    if isinstance(value, dict):
        return (
            _float_or_none(value.get("latitude") or value.get("lat")),
            _float_or_none(value.get("longitude") or value.get("lon") or value.get("lng")),
        )
    if isinstance(value, (list, tuple)) and len(value) >= 2:
        return _float_or_none(value[0]), _float_or_none(value[1])
    text = _text(value)
    if not text:
        return None, None
    numbers = re.findall(r"-?\d+(?:[.,]\d+)?", text)
    if len(numbers) >= 2:
        return _float_or_none(numbers[0]), _float_or_none(numbers[1])
    return None, None


def _nobil_country_code(csmd: dict[str, Any], default_country_code: str = "") -> str:
    for value in (
        csmd.get("Land_code"),
        csmd.get("land_code"),
        csmd.get("Country"),
        csmd.get("country"),
    ):
        code = NOBIL_COUNTRY_BY_LAND_CODE.get(_text(value).upper())
        if code:
            return code
    international_id = _text(csmd.get("International_id")).upper()
    if international_id.startswith("NOR"):
        return "NO"
    if international_id.startswith("SWE"):
        return "SE"
    return default_country_code.upper()


def _nobil_country_code_from_ids(
    *,
    nobil_id: Any = "",
    evse_uid: Any = "",
    default_country_code: str = "",
) -> str:
    nobil_text = _text(nobil_id).upper()
    evse_text = _text(evse_uid).upper()
    for value in (nobil_text, evse_text):
        if value.startswith("NOR") or value.startswith("NO*"):
            return "NO"
        if value.startswith("SWE") or value.startswith("SE*"):
            return "SE"
    return default_country_code.upper()


def _nobil_is_active(csmd: dict[str, Any]) -> bool:
    if "Active" not in csmd:
        return True
    return _text(csmd.get("Active")).casefold() not in {"0", "false", "no", "nei", "nej"}


def _nobil_address(csmd: dict[str, Any]) -> str:
    street = _text(csmd.get("Street"))
    house_number = _text(csmd.get("House_number"))
    if street and house_number and house_number.casefold() not in street.casefold():
        return f"{street} {house_number}".strip()
    return street or house_number


def _nobil_current_type(*source_texts: Any) -> str:
    text = " ".join(_text(value) for value in source_texts).casefold()
    if "dc" in text or "chademo" in text or "combo" in text or "ccs" in text:
        return "DC"
    if "3-phase" in text or "3 phase" in text or "3-fase" in text or "3 fas" in text or "3fas" in text:
        return "AC_3_PHASE"
    if "1-phase" in text or "1 phase" in text or "1-fase" in text or "1 fas" in text or "1fas" in text:
        return "AC_1_PHASE"
    return "AC" if text else ""


def _nobil_power_kw(*source_texts: Any) -> float | None:
    text = " ".join(_text(value) for value in source_texts).replace(",", ".")
    if not text:
        return None
    kw_values = [_float_or_none(match) for match in re.findall(r"(\d+(?:\.\d+)?)\s*kW\b", text, flags=re.IGNORECASE)]
    kw_numbers = [value for value in kw_values if value is not None]
    if kw_numbers:
        return max(kw_numbers)
    watt_values = [_float_or_none(match) for match in re.findall(r"(\d+(?:\.\d+)?)\s*W\b", text, flags=re.IGNORECASE)]
    watt_numbers = [value for value in watt_values if value and value > 1000]
    if watt_numbers:
        return max(watt_numbers) / 1000.0
    voltage_match = re.search(r"(\d+(?:\.\d+)?)\s*V\b", text, flags=re.IGNORECASE)
    amp_match = re.search(r"(\d+(?:\.\d+)?)\s*A\b", text, flags=re.IGNORECASE)
    voltage = _float_or_none(voltage_match.group(1)) if voltage_match else None
    amperage = _float_or_none(amp_match.group(1)) if amp_match else None
    if voltage and amperage:
        current_type = _nobil_current_type(text)
        if current_type == "AC_3_PHASE" and voltage >= 350:
            return round(voltage * amperage * (3**0.5) / 1000.0, 3)
        phase_factor = 3 if current_type == "AC_3_PHASE" else 1
        return round(voltage * amperage * phase_factor / 1000.0, 3)
    return None


def _iter_nobil_rows(
    payload: Any,
    *,
    default_country_code: str,
) -> Iterable[dict[str, Any]]:
    for station in _nobil_station_items(payload):
        csmd = station.get("csmd") if isinstance(station.get("csmd"), dict) else {}
        if not csmd or not _nobil_is_active(csmd):
            continue
        country_code = _nobil_country_code(csmd, default_country_code)
        if default_country_code and country_code != default_country_code.upper():
            continue
        if country_code not in NOBIL_SOURCE_UID_BY_COUNTRY:
            continue
        source_station_id = _text(csmd.get("International_id")) or _text(csmd.get("id"))
        if not source_station_id:
            continue
        attr = station.get("attr") if isinstance(station.get("attr"), dict) else {}
        station_attrs = attr.get("st") if isinstance(attr, dict) else {}
        connector_attrs = attr.get("conn") if isinstance(attr, dict) else {}
        latitude, longitude = _nobil_coordinates(csmd.get("Position") or csmd.get("position"))
        point_count = _int_or_none(csmd.get("Number_charging_points")) or 0
        if point_count < 1:
            point_count = 1
        connection_groups = _nobil_connection_groups(connector_attrs, point_count=point_count)
        country_prefix = country_code.lower()
        source_uid = NOBIL_SOURCE_UID_BY_COUNTRY[country_code]
        provider_uid = NOBIL_PROVIDER_UID_BY_COUNTRY[country_code]
        station_id = f"{country_prefix}:nobil:station:{_safe_id(source_station_id)}"
        for fallback_index, (connector_id, connector_group) in enumerate(connection_groups, start=1):
            normalized_connector_id = connector_id or str(fallback_index)
            connector_types = "|".join(_nobil_attr_texts(connector_group, names={"Connector"}, ids={"4"}))
            capacity_text = "|".join(_nobil_attr_texts(connector_group, names={"Charging capacity"}, ids={"5"}))
            accessibility = "|".join(
                _nobil_attr_texts(connector_group, names={"Accessability", "Accessibility"}, ids={"1"})
            )
            payment_methods = "|".join(_nobil_attr_texts(connector_group, names={"Payment method"}, ids={"19"}))
            power_kw = _nobil_power_kw(capacity_text, connector_types)
            current_type = _nobil_current_type(capacity_text, connector_types)
            evse_uid = _nobil_first_attr_text(
                connector_group,
                names={"EVSE UID"},
                ids={"27"},
                prefer_attrval=True,
            )
            evse_id = _nobil_first_attr_text(
                connector_group,
                names={"EVSE ID"},
                ids={"28"},
                prefer_attrval=True,
            )
            connector_external_id = _nobil_first_attr_text(
                connector_group,
                names={"Connector ID"},
                ids={"29"},
                prefer_attrval=True,
            )
            synthetic_source_evse_id = f"{source_station_id}|point:{normalized_connector_id}"
            source_evse_id = evse_uid or evse_id or synthetic_source_evse_id
            source_evse_alias_ids = [
                alias
                for alias in dict.fromkeys(
                    [
                        source_evse_id,
                        evse_uid,
                        evse_id,
                        connector_external_id,
                        synthetic_source_evse_id,
                    ]
                )
                if alias
            ]
            yield {
                "country_code": country_code,
                "source_uid": source_uid,
                "provider_uid": provider_uid,
                "station_id": station_id,
                "charger_id": f"{country_prefix}:nobil:point:{_safe_id(source_station_id)}:{_safe_id(normalized_connector_id)}",
                "source_station_id": source_station_id,
                "source_evse_id": source_evse_id,
                "source_evse_alias_ids": source_evse_alias_ids,
                "connector_id": normalized_connector_id,
                "connector_types": connector_types,
                "current_type": current_type,
                "max_power_kw": power_kw,
                "operator_name": _text(csmd.get("Owned_by")),
                "station_name": _text(csmd.get("name")),
                "address": _nobil_address(csmd),
                "postal_code": _text(csmd.get("Zipcode")),
                "city": _text(csmd.get("City")),
                "latitude": latitude,
                "longitude": longitude,
                "opening_hours": _nobil_opening_hours(station_attrs),
                "payment_methods": payment_methods,
                "auth_methods": accessibility,
                "green_energy": "",
                "date_updated": _text(csmd.get("Updated")) or _text(csmd.get("Created")),
                "public_bundle_note": "nobil_station_level_point_expansion",
            }


def iter_no_nobil_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    yield from _iter_nobil_rows(payload, default_country_code="NO")


def iter_se_nobil_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    yield from _iter_nobil_rows(payload, default_country_code="SE")


def _parse_json_maybe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text[0] not in "[{":
        return value
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return value


def _nobil_realtime_messages(payload: Any) -> list[dict[str, Any]]:
    payload = _parse_json_maybe(payload)
    if isinstance(payload, dict):
        if any(key in payload for key in ("nobilId", "nobil_id", "evseUId", "evseUid", "status")):
            return [payload]
        messages: list[dict[str, Any]] = []
        for key in ("messages", "events", "items", "data", "payload"):
            if key in payload:
                messages.extend(_nobil_realtime_messages(payload[key]))
        return messages
    if isinstance(payload, list):
        messages = []
        for item in payload:
            messages.extend(_nobil_realtime_messages(item))
        return messages
    return []


def _nobil_realtime_source_observed_at(message: dict[str, Any], fallback_observed_at: str) -> str:
    for key in (
        "source_observed_at",
        "timestamp",
        "lastUpdated",
        "last_updated",
        "updatedAt",
        "updated_at",
        "timeStamp",
    ):
        value = _text(message.get(key))
        if value:
            return value
    return fallback_observed_at


def _iter_nobil_realtime_rows(
    payload: Any,
    *,
    default_country_code: str = "",
) -> Iterable[dict[str, Any]]:
    fallback_observed_at = _text(payload.get("captured_at")) if isinstance(payload, dict) else ""
    for message in _nobil_realtime_messages(payload):
        nobil_id = _text(message.get("nobilId") or message.get("nobil_id"))
        evse_uid = _text(
            message.get("evseUId")
            or message.get("evseUid")
            or message.get("evse_uid")
            or message.get("evseUID")
            or message.get("evse_id")
        )
        source_status = _text(message.get("status")).upper() or "UNKNOWN"
        if not nobil_id or not evse_uid:
            continue
        country_code = _nobil_country_code_from_ids(
            nobil_id=nobil_id,
            evse_uid=evse_uid,
            default_country_code=default_country_code,
        )
        if default_country_code and country_code != default_country_code.upper():
            continue
        if country_code not in NOBIL_REALTIME_SOURCE_UID_BY_COUNTRY:
            continue
        country_prefix = country_code.lower()
        yield {
            "country_code": country_code,
            "source_uid": NOBIL_REALTIME_SOURCE_UID_BY_COUNTRY[country_code],
            "provider_uid": NOBIL_PROVIDER_UID_BY_COUNTRY[country_code],
            "station_id": f"{country_prefix}:nobil:station:{_safe_id(nobil_id)}",
            "charger_id": f"{country_prefix}:nobil:evse:{_safe_id(evse_uid)}",
            "source_station_id": nobil_id,
            "source_evse_id": evse_uid,
            "source_status": source_status,
            "availability_status": NOBIL_REALTIME_STATUS_MAP.get(source_status, "unknown"),
            "source_observed_at": _nobil_realtime_source_observed_at(message, fallback_observed_at),
        }


def iter_no_nobil_realtime_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    yield from _iter_nobil_realtime_rows(payload, default_country_code="NO")


def iter_se_nobil_realtime_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    yield from _iter_nobil_realtime_rows(payload, default_country_code="SE")


LT_OCPI_STATUS_MAP = {
    "AVAILABLE": "free",
    "BLOCKED": "occupied",
    "CHARGING": "occupied",
    "IN_USE": "occupied",
    "OCCUPIED": "occupied",
    "RESERVED": "occupied",
    "FAULTED": "out_of_order",
    "INOPERATIVE": "out_of_order",
    "OUTOFORDER": "out_of_order",
    "OUT_OF_ORDER": "out_of_order",
    "OUTOFSERVICE": "out_of_order",
    "OUT_OF_SERVICE": "out_of_order",
    "REMOVED": "out_of_order",
    "UNAVAILABLE": "out_of_order",
    "UNKNOWN": "unknown",
}


def _lt_legacy_rows(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict) and isinstance(payload.get("pages"), list):
        rows: list[dict[str, Any]] = []
        for page in payload["pages"]:
            response = page.get("response") if isinstance(page, dict) else None
            if isinstance(response, dict):
                rows.extend(response.get("rows") or [])
        return [row for row in rows if isinstance(row, dict)]
    if isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        return [row for row in payload["rows"] if isinstance(row, dict)]
    return []


def _lt_payload_timestamp(payload: Any) -> str:
    return _text(payload.get("timestamp")) if isinstance(payload, dict) else ""


def _lt_ocpi_locations(payload: Any) -> Iterable[dict[str, Any]]:
    if isinstance(payload, list):
        yield from (item for item in payload if isinstance(item, dict))
        return
    if not isinstance(payload, dict):
        return
    data = payload.get("data")
    if isinstance(data, list):
        yield from (item for item in data if isinstance(item, dict))
        return
    if isinstance(data, dict):
        for key in ("locations", "Locations"):
            value = data.get(key)
            if isinstance(value, list):
                yield from (item for item in value if isinstance(item, dict))
                return
    for key in ("locations", "Locations"):
        value = payload.get(key)
        if isinstance(value, list):
            yield from (item for item in value if isinstance(item, dict))
            return


def _lt_ocpi_coordinates(location: dict[str, Any]) -> tuple[float | None, float | None]:
    coordinates = location.get("coordinates") if isinstance(location.get("coordinates"), dict) else {}
    latitude = (
        _float_or_none(coordinates.get("latitude"))
        or _float_or_none(coordinates.get("lat"))
        or _float_or_none(location.get("latitude"))
        or _float_or_none(location.get("lat"))
    )
    longitude = (
        _float_or_none(coordinates.get("longitude"))
        or _float_or_none(coordinates.get("lon"))
        or _float_or_none(coordinates.get("lng"))
        or _float_or_none(location.get("longitude"))
        or _float_or_none(location.get("lon"))
        or _float_or_none(location.get("lng"))
    )
    return latitude, longitude


def _lt_ocpi_station_id(party_id: str, source_station_id: str) -> str:
    if party_id:
        return f"lt:vialietuva:ocpi:{_safe_id(party_id)}:{_safe_id(source_station_id)}"
    return f"lt:vialietuva:ocpi:{_safe_id(source_station_id)}"


def _lt_ocpi_tariff_ids(connectors: Iterable[dict[str, Any]]) -> str:
    values: list[str] = []
    for connector in connectors:
        raw_ids = connector.get("tariff_ids")
        if not isinstance(raw_ids, list):
            raw_ids = [connector.get("tariff_id")]
        for value in raw_ids:
            text = _text(value)
            if text and text not in values:
                values.append(text)
    return "|".join(values)


def _lt_datex_station_id(source_site_id: Any) -> str:
    return f"lt:vialietuva:datex:{_safe_id(source_site_id)}"


def _lt_datex_charger_id(source_evse_id: Any) -> str:
    return f"lt:vialietuva:datex:evse:{_safe_id(source_evse_id)}"


def _lt_datex_source_evse_id(site_id: str, station_id: str, refill_id: str) -> str:
    return f"{site_id}:{station_id}:{refill_id}"


def _lt_datex_reference_id(parent: ET.Element) -> str:
    if _local_name(parent) == "reference":
        return _text(parent.attrib.get("id"))
    for child in parent:
        if _local_name(child) == "reference":
            return _text(child.attrib.get("id"))
    return _text(parent.attrib.get("id"))


def _lt_datex_join_unique(values: Iterable[Any]) -> str:
    seen: dict[str, None] = {}
    for value in values:
        text = _text(value)
        if text:
            seen.setdefault(text, None)
    return "|".join(seen.keys())


def _lt_datex_coordinates(site: ET.Element) -> tuple[float | None, float | None]:
    for container_name in ("coordinatesForDisplay", "pointCoordinates"):
        for element in site.iter():
            if _local_name(element) != container_name:
                continue
            latitude = _float_or_none(_direct_child_text(element, "latitude"))
            longitude = _float_or_none(_direct_child_text(element, "longitude"))
            if latitude is not None and longitude is not None:
                return latitude, longitude
    return None, None


def _lt_datex_address_parts(site: ET.Element) -> dict[str, str]:
    result = {"address": "", "postal_code": "", "city": ""}
    for location_reference in site.iter():
        if _local_name(location_reference) != "locationReference":
            continue
        result["postal_code"] = _first_text(location_reference, "postcode")
        result["city"] = _first_text(location_reference, "city")
        address_lines: list[tuple[int, str]] = []
        for address_line in location_reference.iter():
            if _local_name(address_line) != "addressLine":
                continue
            order_text = _text(address_line.attrib.get("order"))
            order = int(order_text) if order_text.isdigit() else 0
            line = _direct_child_value_text(address_line, "text") or _first_text(address_line, "value")
            if line:
                address_lines.append((order, line))
        result["address"] = " ".join(line for _order, line in sorted(address_lines) if line)
        return result
    return result


def _lt_datex_operator(site: ET.Element) -> tuple[str, str]:
    for child in site:
        if _local_name(child) != "operator":
            continue
        operator_id = _text(child.attrib.get("id"))
        operator_name = _direct_child_value_text(child, "name") or _direct_child_value_text(child, "legalName")
        return operator_id, operator_name or operator_id
    return "", ""


def _lt_datex_current_type(values: Iterable[str]) -> str:
    normalized = " ".join(_text(value).casefold() for value in values)
    if "dc" in normalized or "chademo" in normalized or "combo" in normalized:
        return "DC"
    if "ac" in normalized or "iec62196t2" in normalized or "domestic" in normalized:
        return "AC"
    return ""


def _lt_datex_connector_rows(refill_point: ET.Element) -> list[dict[str, Any]]:
    connectors: list[dict[str, Any]] = []
    for index, connector in enumerate(_direct_children(refill_point, "connector"), start=1):
        connector_type = _direct_child_text(connector, "connectorType")
        charging_mode = _direct_child_text(connector, "chargingMode")
        connector_format = _direct_child_text(connector, "connectorFormat")
        power = _float_or_none(_direct_child_text(connector, "maxPowerAtSocket"))
        if power is not None:
            power = power / 1000.0 if power > 1000 else power
        connector_id = _text(connector.attrib.get("id")) or str(index)
        connectors.append(
            {
                "connector_id": connector_id,
                "connector_type": connector_type,
                "charging_mode": charging_mode,
                "connector_format": connector_format,
                "current_type": _lt_datex_current_type((connector_type, charging_mode)),
                "max_power_kw": power,
            }
        )
    return connectors


def _iter_lt_datex_table_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    text_stream = _text_stream_from_binary_stream(raw_stream, content_encoding=content_encoding)
    publication_time = ""
    seen_evse_ids: set[str] = set()
    for _event, element in ET.iterparse(text_stream, events=("end",)):
        local_name = _local_name(element)
        if local_name == "publicationTime" and not publication_time:
            publication_time = _text(element.text)
            continue
        if local_name != "energyInfrastructureSite":
            continue
        source_site_id = _text(element.attrib.get("id"))
        if not source_site_id:
            element.clear()
            continue
        station_id = _lt_datex_station_id(source_site_id)
        station_name = _direct_child_value_text(element, "name")
        site_lat, site_lon = _lt_datex_coordinates(element)
        address = _lt_datex_address_parts(element)
        operator_id, operator_name = _lt_datex_operator(element)
        for station in element.iter():
            if _local_name(station) != "energyInfrastructureStation":
                continue
            source_station_ref = _text(station.attrib.get("id")) or source_site_id
            auth_methods = _lt_datex_join_unique(
                child.text
                for child in _direct_children(station, "authenticationAndIdentificationMethods")
                if _text(child.text)
            )
            for refill_point in _direct_children(station, "refillPoint"):
                source_refill_id = _text(refill_point.attrib.get("id"))
                if not source_refill_id:
                    continue
                source_evse_id = _lt_datex_source_evse_id(source_site_id, source_station_ref, source_refill_id)
                if source_evse_id in seen_evse_ids:
                    continue
                seen_evse_ids.add(source_evse_id)
                connectors = _lt_datex_connector_rows(refill_point)
                power_values = [
                    connector.get("max_power_kw")
                    for connector in connectors
                    if connector.get("max_power_kw") is not None
                ]
                yield {
                    "country_code": "LT",
                    "source_uid": LT_EV_LOCATIONS_SOURCE_UID,
                    "provider_uid": "lt_vialietuva_datex",
                    "station_id": station_id,
                    "charger_id": _lt_datex_charger_id(source_evse_id),
                    "source_station_id": source_site_id,
                    "source_station_ref": source_station_ref,
                    "source_evse_id": source_evse_id,
                    "connector_id": _lt_datex_join_unique(connector.get("connector_id") for connector in connectors),
                    "connector_types": _lt_datex_join_unique(connector.get("connector_type") for connector in connectors),
                    "connector_formats": _lt_datex_join_unique(connector.get("connector_format") for connector in connectors),
                    "current_type": _lt_datex_join_unique(connector.get("current_type") for connector in connectors),
                    "max_power_kw": max(power_values) if power_values else None,
                    "connector_count": len(connectors) or 1,
                    "operator_name": operator_name or operator_id,
                    "station_name": station_name or source_site_id,
                    "address": address["address"],
                    "postal_code": address["postal_code"],
                    "city": address["city"],
                    "latitude": site_lat,
                    "longitude": site_lon,
                    "auth_methods": auth_methods,
                    "date_updated": publication_time,
                    "raw_static": _compact_json(
                        {
                            "site_id": source_site_id,
                            "station_id": source_station_ref,
                            "refill_point_id": source_refill_id,
                        }
                    ),
                }
        element.clear()


def _lt_datex_status_from_source(source_status: str, is_available: str) -> str:
    normalized = _text(source_status).upper() or "UNKNOWN"
    if _text(is_available).casefold() == "false" and normalized == "UNKNOWN":
        return "out_of_order"
    return LT_OCPI_STATUS_MAP.get(normalized, "unknown")


def _iter_lt_datex_status_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    text_stream = _text_stream_from_binary_stream(raw_stream, content_encoding=content_encoding)
    publication_time = ""
    for _event, element in ET.iterparse(text_stream, events=("end",)):
        local_name = _local_name(element)
        if local_name == "publicationTime" and not publication_time:
            publication_time = _text(element.text)
            continue
        if local_name != "energyInfrastructureSiteStatus":
            continue
        source_site_id = _lt_datex_reference_id(element)
        if not source_site_id:
            element.clear()
            continue
        station_id = _lt_datex_station_id(source_site_id)
        site_updated = _direct_child_text(element, "lastUpdated")
        for station_status in element.iter():
            if _local_name(station_status) != "energyInfrastructureStationStatus":
                continue
            source_station_ref = _lt_datex_reference_id(station_status) or source_site_id
            station_available = _direct_child_text(station_status, "isAvailable")
            station_updated = _direct_child_text(station_status, "lastUpdated")
            for refill_status in _direct_children(station_status, "refillPointStatus"):
                source_refill_id = _lt_datex_reference_id(refill_status)
                if not source_refill_id:
                    continue
                source_evse_id = _lt_datex_source_evse_id(source_site_id, source_station_ref, source_refill_id)
                source_status = _direct_child_text(refill_status, "status").upper() or "UNKNOWN"
                observed_at = (
                    _direct_child_text(refill_status, "lastUpdated")
                    or station_updated
                    or site_updated
                    or publication_time
                )
                yield {
                    "country_code": "LT",
                    "source_uid": LT_EV_STATUS_SOURCE_UID,
                    "provider_uid": "lt_vialietuva_datex",
                    "station_id": station_id,
                    "charger_id": _lt_datex_charger_id(source_evse_id),
                    "source_station_id": source_site_id,
                    "source_station_ref": source_station_ref,
                    "source_evse_id": source_evse_id,
                    "source_status": source_status,
                    "availability_status": _lt_datex_status_from_source(source_status, station_available),
                    "source_observed_at": observed_at,
                }
        element.clear()


def _lt_stream_looks_like_xml(raw_stream: io.BufferedIOBase) -> bool:
    if hasattr(raw_stream, "peek"):
        return bytes(raw_stream.peek(128)).lstrip().startswith(b"<")
    return False


def _iter_lt_legacy_rows(payload: Any) -> Iterable[dict[str, Any]]:
    seen_source_evses: set[str] = set()
    for location in _lt_legacy_rows(payload):
        source_station_id = _text(location.get("id"))
        if not source_station_id:
            continue
        station_id = f"lt:lakd:loc:{_safe_id(source_station_id)}"
        location_data = location.get("l") if isinstance(location.get("l"), dict) else location
        latitude = _float_or_none(location_data.get("x"))
        longitude = _float_or_none(location_data.get("y"))
        operator = location.get("o") if isinstance(location.get("o"), dict) else {}
        operator_name = _text(operator.get("name")) or _text(location.get("n"))
        for evse in location.get("e") or []:
            if not isinstance(evse, dict):
                continue
            source_evse_id = _text(evse.get("eid")) or _text(evse.get("id"))
            if not source_evse_id:
                continue
            evse_key = source_evse_id.casefold()
            if evse_key in seen_source_evses:
                continue
            seen_source_evses.add(evse_key)
            connectors = [item for item in evse.get("c") or [] if isinstance(item, dict)]
            row = {
                "country_code": "LT",
                "source_uid": LT_LEGACY_EV_LOCATIONS_SOURCE_UID,
                "provider_uid": "lt_ev_lakd",
                "station_id": station_id,
                "charger_id": f"lt:lakd:evse:{_safe_id(source_evse_id)}",
                "source_station_id": source_station_id,
                "source_evse_id": source_evse_id,
                "connector_id": "|".join(dict.fromkeys(_text(item.get("id")) for item in connectors if _text(item.get("id")))),
                "connector_types": "|".join(dict.fromkeys(_text(item.get("sdr")) for item in connectors if _text(item.get("sdr")))),
                "current_type": "",
                "max_power_kw": max((_float_or_none(item.get("kw")) for item in connectors if _float_or_none(item.get("kw")) is not None), default=None),
                "operator_name": operator_name,
                "station_name": _text(location.get("n")),
                "address": _text(location.get("adr")),
                "postal_code": "",
                "city": _text(location.get("c")),
                "latitude": latitude,
                "longitude": longitude,
                "date_updated": _text(location.get("lu")),
                "opening_hours": _lt_opening_hours(location.get("ot")),
                "helpdesk_phone": _text(location.get("tel")),
            }
            row.update(_merge_price_field_rows([_lt_price_fields(item.get("price")) for item in connectors]))
            yield row


def _iter_lt_ocpi_rows(
    payload: Any,
    *,
    source_uid: str,
    include_dynamic: bool,
) -> Iterable[dict[str, Any]]:
    seen_source_evses: set[str] = set()
    fallback_observed_at = _lt_payload_timestamp(payload)
    for location in _lt_ocpi_locations(payload):
        country_code = _text(location.get("country_code") or location.get("country")).upper()
        if country_code and country_code not in {"LT", "LTU"}:
            continue
        source_station_id = _text(location.get("id"))
        if not source_station_id:
            continue
        party_id = _text(location.get("party_id"))
        station_id = _lt_ocpi_station_id(party_id, source_station_id)
        operator = location.get("operator") if isinstance(location.get("operator"), dict) else {}
        owner = location.get("owner") if isinstance(location.get("owner"), dict) else {}
        operator_name = _text(operator.get("name")) or _text(owner.get("name")) or _text(location.get("name"))
        latitude, longitude = _lt_ocpi_coordinates(location)
        evses = [item for item in location.get("evses") or [] if isinstance(item, dict)]
        for evse in evses:
            if not isinstance(evse, dict):
                continue
            source_evse_id = _text(evse.get("evse_id")) or _text(evse.get("uid"))
            if not source_evse_id:
                continue
            evse_key = source_evse_id.casefold()
            if evse_key in seen_source_evses:
                continue
            seen_source_evses.add(evse_key)
            connectors = [item for item in evse.get("connectors") or [] if isinstance(item, dict)]
            source_status = _text(evse.get("status")).upper() or "UNKNOWN"
            row = {
                "country_code": "LT",
                "source_uid": source_uid,
                "provider_uid": "lt_vialietuva_ocpi",
                "station_id": station_id,
                "charger_id": f"lt:vialietuva:ocpi:evse:{_safe_id(source_evse_id)}",
                "source_station_id": source_station_id,
                "source_evse_id": source_evse_id,
                "connector_id": "|".join(dict.fromkeys(_text(item.get("id")) for item in connectors if _text(item.get("id")))),
                "connector_types": "|".join(dict.fromkeys(_text(item.get("standard")) for item in connectors if _text(item.get("standard")))),
                "current_type": "|".join(dict.fromkeys(_text(item.get("power_type")).upper() for item in connectors if _text(item.get("power_type")))),
                "max_power_kw": _max_power_from_connectors(connectors),
                "operator_name": operator_name,
                "station_name": _text(location.get("name")) or _text(location.get("address")),
                "address": _text(location.get("address")),
                "postal_code": _text(location.get("postal_code")),
                "city": _text(location.get("city")),
                "latitude": latitude,
                "longitude": longitude,
                "date_updated": _text(evse.get("last_updated")) or _text(location.get("last_updated")),
                "opening_hours": _lt_opening_hours(location.get("opening_times")),
                "helpdesk_phone": _text(location.get("help_phone")),
                "tariff_ids": _lt_ocpi_tariff_ids(connectors),
            }
            if include_dynamic:
                row.update(
                    {
                        "source_status": source_status,
                        "availability_status": LT_OCPI_STATUS_MAP.get(source_status, "unknown"),
                        "source_observed_at": (
                            _text(evse.get("last_updated"))
                            or _text(location.get("last_updated"))
                            or fallback_observed_at
                        ),
                    }
                )
            else:
                row["raw_static"] = _compact_json(
                    {
                        "location": {key: value for key, value in location.items() if key != "evses"},
                        "evse": {key: value for key, value in evse.items() if key not in {"connectors", "status"}},
                        "connectors": connectors,
                    }
                )
            yield row


def iter_lt_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    if _lt_stream_looks_like_xml(raw_stream):
        yield from _iter_lt_datex_table_rows_from_binary_stream(raw_stream, content_encoding=content_encoding)
        return
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    legacy_rows = _lt_legacy_rows(payload)
    if legacy_rows:
        yield from _iter_lt_legacy_rows({"rows": legacy_rows})
        return
    yield from _iter_lt_ocpi_rows(
        payload,
        source_uid=LT_EV_LOCATIONS_SOURCE_UID,
        include_dynamic=False,
    )


def iter_lt_dynamic_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    if _lt_stream_looks_like_xml(raw_stream):
        yield from _iter_lt_datex_status_rows_from_binary_stream(raw_stream, content_encoding=content_encoding)
        return
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    yield from _iter_lt_ocpi_rows(
        payload,
        source_uid=LT_EV_STATUS_SOURCE_UID,
        include_dynamic=True,
    )


def _lu_power_and_connector(description: str) -> tuple[int, float | None, str]:
    count = 1
    count_match = re.search(r"(\d+)\s+connectors?", description, flags=re.IGNORECASE)
    if count_match:
        count = int(count_match.group(1))
    power_match = re.search(r"(\d+(?:[.,]\d+)?)\s*kW", description, flags=re.IGNORECASE)
    connector = ""
    connector_match = re.search(r"and\s+(.+?)\s+connector", description, flags=re.IGNORECASE)
    if connector_match:
        connector = connector_match.group(1).strip()
    return count, _float_or_none(power_match.group(1) if power_match else ""), connector


def iter_lu_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    for feature in payload.get("features") or []:
        if not isinstance(feature, dict):
            continue
        props = feature.get("properties") or {}
        geometry = feature.get("geometry") or {}
        coords = geometry.get("coordinates") or []
        if len(coords) < 2:
            continue
        source_station_id = _text(props.get("inspireid_identifier_localid")) or _text(feature.get("id"))
        if not source_station_id:
            continue
        description = _text(props.get("gml_description"))
        _connector_count, max_power_kw, connector_type = _lu_power_and_connector(description)
        station_id = f"lu:data-public:node:{_safe_id(source_station_id)}"
        source_evse_id = f"{source_station_id}|station-placeholder"
        yield {
            "country_code": "LU",
            "source_uid": LU_CHARGING_STATIONS_SOURCE_UID,
            "provider_uid": "lu_data_public",
            "station_id": station_id,
            "charger_id": f"lu:data-public:station-placeholder:{_safe_id(source_station_id)}",
            "source_station_id": source_station_id,
            "source_evse_id": source_evse_id,
            "connector_id": "",
            "connector_types": connector_type,
            "current_type": "",
            "max_power_kw": max_power_kw,
            "operator_name": "",
            "station_name": description.split("|", 1)[0].strip(),
            "address": "",
            "postal_code": "",
            "city": "",
            "latitude": _float_or_none(coords[1]),
            "longitude": _float_or_none(coords[0]),
        }


def iter_mt_rows_from_binary_stream(
    raw_stream: io.BufferedIOBase,
    *,
    content_encoding: str = "",
) -> Iterable[dict[str, Any]]:
    payload = _json_from_binary_stream(raw_stream, content_encoding=content_encoding)
    for feature in payload.get("features") or []:
        if not isinstance(feature, dict):
            continue
        props = feature.get("properties") or {}
        geometry = feature.get("geometry") or {}
        coords = geometry.get("coordinates") or []
        source_station_id = _text(props.get("Ref")) or _text(props.get("OBJECTID")) or _text(feature.get("id"))
        if not source_station_id or len(coords) < 2:
            continue
        charger_type = _text(props.get("Charger_Type"))
        max_power_kw = 50.0 if charger_type.casefold() == "fast" else None
        station_id = f"mt:egis:{_safe_id(source_station_id)}"
        yield {
            "country_code": "MT",
            "source_uid": MT_CHARGING_POINTS_SOURCE_UID,
            "provider_uid": "mt_transport_geoservices",
            "station_id": station_id,
            "charger_id": f"mt:egis:point:{_safe_id(source_station_id)}",
            "source_station_id": source_station_id,
            "source_evse_id": source_station_id,
            "connector_id": "",
            "connector_types": charger_type,
            "current_type": "DC" if charger_type.casefold() == "fast" else "",
            "max_power_kw": max_power_kw,
            "operator_name": "Transport Malta",
            "station_name": f"{_text(props.get('Locality'))} {_text(props.get('Road_Area'))}".strip(),
            "address": _text(props.get("Road_Area")),
            "postal_code": "",
            "city": _text(props.get("Locality")),
            "latitude": _float_or_none(coords[1]),
            "longitude": _float_or_none(coords[0]),
            "date_updated": _text(props.get("State")),
        }
