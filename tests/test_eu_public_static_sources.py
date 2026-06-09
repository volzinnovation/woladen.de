from __future__ import annotations

import hashlib
import io
import json
import urllib.error
import zipfile
from dataclasses import replace
from email.message import Message

import pytest
from commercial_backend.config import AppConfig
from openpyxl import Workbook

from commercial_backend.eu_public_static import (
    LT_EV_LOCATIONS_SOURCE_UID,
    LT_EV_STATUS_SOURCE_UID,
    PUBLIC_STATIC_SOURCES,
    iter_cy_rows_from_binary_stream,
    iter_cz_rows_from_binary_stream,
    iter_es_rows_from_binary_stream,
    iter_gr_dynamic_rows_from_binary_stream,
    iter_gr_rows_from_binary_stream,
    iter_lt_dynamic_rows_from_binary_stream,
    iter_lt_rows_from_binary_stream,
    iter_lu_rows_from_binary_stream,
    iter_mt_rows_from_binary_stream,
    iter_no_nobil_rows_from_binary_stream,
    iter_no_nobil_realtime_rows_from_binary_stream,
    iter_se_nobil_rows_from_binary_stream,
    iter_se_nobil_realtime_rows_from_binary_stream,
)
from scripts import commercial_fetch_eu_public_static as fetch_eu


def _stream(payload: bytes) -> io.BufferedReader:
    return io.BufferedReader(io.BytesIO(payload))


def _json_stream(payload: dict) -> io.BufferedReader:
    return _stream(json.dumps(payload).encode("utf-8"))


def _zip_json_stream(payload: dict, name: str = "payload.json") -> io.BufferedReader:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(name, json.dumps(payload).encode("utf-8"))
    return _stream(buffer.getvalue())


def test_cy_groups_same_provider_coordinate_as_one_station():
    payload = b"""
    <payload>
      <chargingPoint>
        <chargingPointIdentification>CY*ABC*E1</chargingPointIdentification>
        <chargingPointOperatorLegalName>Provider A</chargingPointOperatorLegalName>
        <latitude>35.10000001</latitude>
        <longitude>33.20000001</longitude>
        <connectorTypes><connectorType>IEC_62196_T2</connectorType></connectorTypes>
        <maximumPower>22</maximumPower>
      </chargingPoint>
      <chargingPoint>
        <chargingPointIdentification>CY*ABC*E2</chargingPointIdentification>
        <chargingPointOperatorLegalName>Provider A</chargingPointOperatorLegalName>
        <latitude>35.10000001</latitude>
        <longitude>33.20000001</longitude>
        <connectorTypes><connectorType>IEC_62196_T2_COMBO</connectorType></connectorTypes>
        <maximumPower>50</maximumPower>
      </chargingPoint>
    </payload>
    """

    rows = list(iter_cy_rows_from_binary_stream(_stream(payload)))

    assert len(rows) == 2
    assert {row["station_id"] for row in rows} == {"cy:coord:provider-a:35.1000000:33.2000000"}
    assert {row["charger_id"] for row in rows} == {"cy:datex:evse:cy*abc*e1", "cy:datex:evse:cy*abc*e2"}


def test_cz_uses_source_charging_point_count_not_raw_connector_count():
    workbook = Workbook()
    sheet = workbook.active
    row = 8
    sheet.cell(row=row, column=2, value="Main Street 1")
    sheet.cell(row=row, column=3, value="parking")
    sheet.cell(row=row, column=4, value="11000")
    sheet.cell(row=row, column=5, value="Praha")
    sheet.cell(row=row, column=7, value=50.087)
    sheet.cell(row=row, column=8, value=14.421)
    sheet.cell(row=row, column=9, value="CPO")
    sheet.cell(row=row, column=10, value="Operator")
    sheet.cell(row=row, column=12, value=1)
    sheet.cell(row=row, column=13, value="DC")
    sheet.cell(row=row, column=14, value=150)
    sheet.cell(row=row, column=15, value="CCS")
    sheet.cell(row=row, column=18, value="AC")
    sheet.cell(row=row, column=19, value=22)
    sheet.cell(row=row, column=20, value="Type 2")
    sheet.cell(row=row, column=28, value="2026-03-31")
    buffer = io.BytesIO()
    workbook.save(buffer)

    rows = list(iter_cz_rows_from_binary_stream(_stream(buffer.getvalue())))

    assert len(rows) == 1
    assert rows[0]["source_evse_id"].endswith("|point:1")
    assert rows[0]["connector_id"] == "1;2"
    assert rows[0]["connector_types"] == "CCS|Type 2"
    assert rows[0]["max_power_kw"] == 150


def test_es_dgt_datex_rows_use_dgt_station_and_official_evse_ids():
    payload = b"""
    <payload>
      <publicationTime>2026-05-11T10:18:25.927+02:00</publicationTime>
      <energyInfrastructureSite id="SITE-1">
        <name><values><value>Madrid Central Padel</value></values></name>
        <lastUpdated>2026-05-07T15:20:38.000+02:00</lastUpdated>
        <operatingHours id="24/7"/>
        <locationReference>
          <address>
            <postcode>28052</postcode>
            <addressLine><text><values><value>Direccion: Calle Boyer 22</value></values></text></addressLine>
            <addressLine><text><values><value>Municipio: Madrid</value></values></text></addressLine>
          </address>
          <coordinatesForDisplay>
            <latitude>40.403988</latitude>
            <longitude>-3.5875702</longitude>
          </coordinatesForDisplay>
        </locationReference>
        <operator id="ES*AEQ"><name><values><value>QWELLO Espana SL</value></values></name></operator>
        <energyInfrastructureStation id="SITE-1_1">
          <authenticationAndIdentificationMethods>rfid</authenticationAndIdentificationMethods>
          <authenticationAndIdentificationMethods>creditCard</authenticationAndIdentificationMethods>
          <refillPoint id="POINT-1">
            <name><values><value>ES*AEQ*ESAEQEM5PP812</value></values></name>
            <connector>
              <connectorType>iec62196T2COMBO</connectorType>
              <chargingMode>mode4DC</chargingMode>
              <connectorFormat>cable</connectorFormat>
              <maxPowerAtSocket>300000.0</maxPowerAtSocket>
            </connector>
            <connector>
              <connectorType>chademo</connectorType>
              <chargingMode>mode4DC</chargingMode>
              <maxPowerAtSocket>80000.0</maxPowerAtSocket>
            </connector>
          </refillPoint>
        </energyInfrastructureStation>
      </energyInfrastructureSite>
    </payload>
    """

    rows = list(iter_es_rows_from_binary_stream(_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["station_id"] == "es:dgt:site:site-1"
    assert rows[0]["charger_id"] == "es:dgt:evse:es*aeq*esaeqem5pp812"
    assert rows[0]["source_evse_id"] == "ES*AEQ*ESAEQEM5PP812"
    assert rows[0]["provider_uid"] == "es_dgt_electrolineras"
    assert rows[0]["address"] == "Calle Boyer 22"
    assert rows[0]["city"] == "Madrid"
    assert rows[0]["connector_types"] == "iec62196T2COMBO|chademo"
    assert rows[0]["current_type"] == "DC"
    assert rows[0]["max_power_kw"] == 300
    assert rows[0]["auth_methods"] == "rfid|creditCard"


def test_gr_emits_evse_details_and_location_placeholder_for_missing_details():
    payload = {
        "details": {
            "loc-1": {
                "Loc": {
                    "id": "loc-1",
                    "name": "Detailed Location",
                    "address": "Street 1",
                    "city": "Athens",
                    "coordinates": {"latitude": 37.98, "longitude": 23.73},
                    "operator": {"name": "Provider GR"},
                    "evses": [
                        {
                            "evse_id": "GR*ABC*E1",
                            "connectors": [
                                {
                                    "id": "1",
                                    "standard": "IEC_62196_T2_COMBO",
                                    "power_type": "DC",
                                    "max_electric_power": 50000,
                                }
                            ],
                        }
                    ],
                }
            }
        },
        "locations": {
            "features": [
                {"properties": {"location_id": "loc-1"}, "geometry": {"coordinates": [37.98, 23.73]}},
                {
                    "properties": {
                        "location_id": "loc-2",
                        "Provider": "Fallback Provider",
                        "LocationName": "Fallback Location",
                    },
                    "geometry": {"coordinates": [38.0, 24.0]},
                },
            ]
        },
    }

    rows = list(iter_gr_rows_from_binary_stream(_json_stream(payload)))

    assert [row["charger_id"] for row in rows] == [
        "gr:electrokinisi:evse:gr*abc*e1",
        "gr:electrokinisi:location-placeholder:loc-2",
    ]
    assert rows[0]["max_power_kw"] == 50
    assert rows[1]["public_bundle_note"] == "location_only_placeholder_due_detail_api_rate_limit"


def test_gr_idro_static_zip_emits_official_evse_rows():
    payload = {
        "Locations": [
            {
                "country_code": "GR",
                "party_id": "PPC",
                "id": "GR-PPC-LOC-1",
                "publish": True,
                "name": "Town Hall",
                "address": "Main 1",
                "city": "Athens",
                "postal_code": "10000",
                "coordinates": {"latitude": "37.98", "longitude": "23.73"},
                "operator": {"name": "DEI Blue"},
                "opening_times": {"twentyfourseven": True},
                "energy_mix": {"is_green_energy": True},
                "evses": [
                    {
                        "uid": "GR-PPC-E1",
                        "evse_id": "GR*PPC*E1",
                        "status": "AVAILABLE",
                        "supports_roaming": True,
                        "capabilities": ["REMOTE_START_STOP_CAPABLE"],
                        "connectors": [
                            {
                                "id": "1",
                                "standard": "IEC_62196_T2",
                                "format": "SOCKET",
                                "power_type": "AC_3_PHASE",
                                "max_voltage": 230,
                                "max_amperage": 32,
                                "max_electric_power": 22000,
                            }
                        ],
                        "last_updated": "2026-05-11T01:04:42",
                    }
                ],
                "last_updated": "2026-05-11T01:04:43",
            }
        ],
        "status": "ok",
        "statusDesc": "success",
    }

    rows = list(iter_gr_rows_from_binary_stream(_zip_json_stream(payload, "GR.IDRO.static.data.latest.json")))

    assert len(rows) == 1
    assert rows[0]["station_id"] == "gr:electrokinisi:loc:gr-ppc-loc-1"
    assert rows[0]["charger_id"] == "gr:electrokinisi:evse:gr*ppc*e1"
    assert rows[0]["source_evse_id"] == "GR*PPC*E1"
    assert rows[0]["max_power_kw"] == 22
    assert rows[0]["payment_methods"] == "roaming"
    assert rows[0]["auth_methods"] == "REMOTE_START_STOP_CAPABLE"
    assert rows[0]["green_energy"] == "true"


def test_gr_idro_dynamic_zip_maps_status_rows():
    payload = {
        "Locations": [
            {
                "country_code": "GR",
                "party_id": "PPC",
                "id": "GR-PPC-LOC-1",
                "name": "Town Hall",
                "evses": [
                    {
                        "uid": "GR-PPC-E1",
                        "evse_id": "GR*PPC*E1",
                        "status": "CHARGING",
                        "connectors": [{"id": "1"}],
                        "last_updated": "2026-05-11T12:40:20",
                    }
                ],
                "last_updated": "2026-05-11T12:40:19",
            }
        ],
        "status": "ok",
        "statusDesc": "success",
    }

    rows = list(iter_gr_dynamic_rows_from_binary_stream(_zip_json_stream(payload, "GR.IDRO.dynamic.data.latest.json")))

    assert rows == [
        {
            "country_code": "GR",
            "source_uid": "gr_idro_dynamic_json",
            "provider_uid": "gr_electrokinisi",
            "station_id": "gr:electrokinisi:loc:gr-ppc-loc-1",
            "charger_id": "gr:electrokinisi:evse:gr*ppc*e1",
            "source_station_id": "GR-PPC-LOC-1",
            "source_evse_id": "GR*PPC*E1",
            "source_status": "CHARGING",
            "availability_status": "occupied",
            "source_observed_at": "2026-05-11T12:40:20",
            "connector_id": "1",
        }
    ]


def test_invalid_gr_dynamic_zip_is_quarantined_before_queueing(tmp_path):
    config = replace(AppConfig(), raw_payload_dir=tmp_path / "raw")
    temp_path = config.raw_payload_dir / "_incoming" / "not-a-zip.json.zip"
    temp_path.parent.mkdir(parents=True)
    payload = b"not a zip"
    temp_path.write_bytes(payload)
    fetched = fetch_eu.FetchResult(
        payload_path=temp_path,
        payload_sha256="c" * 64,
        byte_length=len(payload),
        content_type="application/zip",
        content_encoding="",
        final_url=PUBLIC_STATIC_SOURCES["gr-dynamic"].url,
    )

    try:
        fetch_eu._validate_gr_payload(
            config=config,
            spec=PUBLIC_STATIC_SOURCES["gr-dynamic"],
            fetched=fetched,
        )
    except RuntimeError as exc:
        assert "invalid_eu_public_payload:gr-dynamic" in str(exc)
    else:
        raise AssertionError("invalid GR ZIP should fail validation")

    assert not temp_path.exists()
    invalid_files = list((config.raw_payload_dir / "_invalid").rglob("*.json.zip"))
    assert len(invalid_files) == 1
    manifest_text = "\n".join(
        path.read_text(encoding="utf-8")
        for path in (config.raw_payload_dir / "_invalid").rglob("invalid_payloads.ndjson")
    )
    assert "gr_idro_dynamic_json" in manifest_text


def test_no_nobil_datadump_expands_station_level_point_count():
    payload = {
        "Provider": "NOBIL.no",
        "Rights": "Creative Commons Attribution 4.0 International License",
        "apiver": "3",
        "chargerstations": [
            {
                "csmd": {
                    "id": 41,
                    "name": "Recharge Filipstad",
                    "Active": "1",
                    "Street": "Harbor Street",
                    "House_number": "2",
                    "Zipcode": "0250",
                    "City": "Oslo",
                    "Owned_by": "Recharge",
                    "Number_charging_points": "2",
                    "Position": "(59.905,10.721)",
                    "Land_code": "NOR",
                    "International_id": "NOR_00041",
                    "Updated": "2026-05-11 12:00:00",
                },
                "attr": {
                    "st": {"24": {"attrname": "Open 24h", "trans": "Yes"}},
                    "conn": {
                        "1": {"attrname": "Accessability", "trans": "Open"},
                        "4": {"attrname": "Connector", "trans": "Type 2"},
                        "5": {"attrname": "Charging capacity", "trans": "230V 1-phase max 16A"},
                    },
                },
            }
        ],
    }

    rows = list(iter_no_nobil_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 2
    assert {row["charger_id"] for row in rows} == {
        "no:nobil:point:nor_00041:1",
        "no:nobil:point:nor_00041:2",
    }
    assert rows[0]["station_id"] == "no:nobil:station:nor_00041"
    assert rows[0]["source_uid"] == "no_nobil_static_datadump"
    assert rows[0]["provider_uid"] == "no_nobil"
    assert rows[0]["source_evse_id"] == "NOR_00041|point:1"
    assert rows[0]["connector_types"] == "Type 2"
    assert rows[0]["current_type"] == "AC_1_PHASE"
    assert rows[0]["max_power_kw"] == 3.68
    assert rows[0]["opening_hours"] == "24/7"
    assert rows[0]["auth_methods"] == "Open"


def test_se_nobil_datadump_uses_sweden_source_uid_and_skips_inactive_rows():
    payload = {
        "Provider": "NOBIL.no",
        "Rights": "Creative Commons Attribution 4.0 International License",
        "apiver": "3",
        "chargerstations": [
            {
                "csmd": {
                    "id": 40442,
                    "name": "Wasby Golf",
                    "Active": "1",
                    "Street": "Golfvagen",
                    "Zipcode": "194 98",
                    "City": "Upplands Vasby",
                    "Owned_by": "CPO SE",
                    "Number_charging_points": "1",
                    "Position": "59.521,17.912",
                    "Land_code": "SWE",
                    "International_id": "SWE_40442",
                },
                "attr": {
                    "conn": {
                        "4": {"attrname": "Connector", "trans": "CCS"},
                        "5": {"attrname": "Charging capacity", "trans": "150 kW DC"},
                    }
                },
            },
            {"csmd": {"id": 1, "Active": "0", "Land_code": "SWE", "International_id": "SWE_00001"}},
        ],
    }

    rows = list(iter_se_nobil_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["country_code"] == "SE"
    assert rows[0]["source_uid"] == "se_nobil_static_datadump"
    assert rows[0]["provider_uid"] == "se_nobil"
    assert rows[0]["station_id"] == "se:nobil:station:swe_40442"
    assert rows[0]["charger_id"] == "se:nobil:point:swe_40442:1"
    assert rows[0]["connector_types"] == "CCS"
    assert rows[0]["current_type"] == "DC"
    assert rows[0]["max_power_kw"] == 150


def test_nobil_datadump_prefers_evse_uid_and_keeps_match_aliases():
    payload = {
        "chargerstations": [
            {
                "csmd": {
                    "id": 74759,
                    "Active": "1",
                    "Number_charging_points": "1",
                    "Position": "(59.910327,10.720555)",
                    "Land_code": "NOR",
                    "International_id": "NOR_74759",
                },
                "attr": {
                    "conn": {
                        "1": {
                            "4": {"attrname": "Connector", "trans": "CCS/Combo"},
                            "5": {"attrname": "Charging capacity", "trans": "150 kW DC"},
                            "27": {"attrname": "EVSE UID", "trans": "EVSE UID", "attrval": "10327"},
                            "28": {"attrname": "EVSE ID", "trans": "EVSE ID", "attrval": "NO*CHA*E4731*B"},
                            "29": {"attrname": "Connector ID", "trans": "Connector ID", "attrval": "10327"},
                        }
                    }
                },
            }
        ]
    }

    rows = list(iter_no_nobil_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["source_evse_id"] == "10327"
    assert rows[0]["source_evse_alias_ids"] == [
        "10327",
        "NO*CHA*E4731*B",
        "NOR_74759|point:1",
    ]


def test_nobil_realtime_parser_maps_no_and_se_status_messages():
    payload = {
        "captured_at": "2026-05-12T10:00:00+00:00",
        "messages": [
            {"nobilId": "NOR_74759", "evseUId": "10327", "status": "AVAILABLE"},
            {"nobilId": "SWE_40442", "evseUId": "0b21c31f-2eb3-493a-57b6-08dd363a7b6b", "status": "CHARGING"},
        ],
    }

    no_rows = list(iter_no_nobil_realtime_rows_from_binary_stream(_json_stream(payload)))
    se_rows = list(iter_se_nobil_realtime_rows_from_binary_stream(_json_stream(payload)))

    assert [row["source_uid"] for row in no_rows] == ["no_nobil_realtime"]
    assert no_rows[0]["provider_uid"] == "no_nobil"
    assert no_rows[0]["station_id"] == "no:nobil:station:nor_74759"
    assert no_rows[0]["source_evse_id"] == "10327"
    assert no_rows[0]["availability_status"] == "free"
    assert no_rows[0]["source_observed_at"] == "2026-05-12T10:00:00+00:00"
    assert [row["source_uid"] for row in se_rows] == ["se_nobil_realtime"]
    assert se_rows[0]["provider_uid"] == "se_nobil"
    assert se_rows[0]["availability_status"] == "occupied"


def test_lt_duplicate_evse_aliases_are_suppressed_for_strict_charger_id_join_key():
    evse = {"eid": "LT*ABC*E1", "c": [{"id": "1", "sdr": "CCS", "kw": 150, "price": "0.40 €/kWh"}]}
    payload = {
        "pages": [
            {
                "response": {
                    "rows": [
                        {"id": "loc-a", "n": "A", "l": {"x": 55.0, "y": 25.0}, "tel": 37062247472, "e": [evse]},
                        {"id": "loc-b", "n": "B", "l": {"x": 55.1, "y": 25.1}, "e": [dict(evse, eid="lt*abc*e1")]},
                    ]
                }
            }
        ]
    }

    rows = list(iter_lt_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["charger_id"] == "lt:lakd:evse:lt*abc*e1"
    assert rows[0]["source_station_id"] == "loc-a"
    assert rows[0]["price_display"] == "0,40 €/kWh"
    assert rows[0]["price_energy_eur_kwh_min"] == "0.4"
    assert rows[0]["helpdesk_phone"] == "37062247472"


def test_lt_via_lietuva_ocpi_static_parser_uses_current_public_endpoint_shape():
    payload = {
        "timestamp": "2026-06-01T07:09:02Z",
        "status_code": 1000,
        "data": [
            {
                "country_code": "LT",
                "party_id": "IBG",
                "id": 257,
                "name": "Z135 | Inbalance grid",
                "address": "Zalgirio g. 135",
                "city": "Vilnius",
                "coordinates": {"latitude": "54.7049540", "longitude": "25.2724750"},
                "operator": {"name": "In Balance grid, UAB"},
                "help_phone": "+37062247472",
                "last_updated": "2026-06-01T07:59:53",
                "evses": [
                    {
                        "uid": 12029,
                        "evse_id": "IBG-P-TYJE-B",
                        "status": "CHARGING",
                        "connectors": [
                            {
                                "id": 650305,
                                "standard": "IEC_62196_T2",
                                "power_type": "AC_1_PHASE",
                                "max_electric_power": 22000,
                                "tariff_ids": ["418c326c-089e-11f0-a250-0a58a9feac0d"],
                            }
                        ],
                        "last_updated": "2026-06-01T07:59:53",
                    }
                ],
            }
        ],
    }

    rows = list(iter_lt_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["source_uid"] == LT_EV_LOCATIONS_SOURCE_UID
    assert rows[0]["station_id"] == "lt:vialietuva:ocpi:ibg:257"
    assert rows[0]["charger_id"] == "lt:vialietuva:ocpi:evse:ibg-p-tyje-b"
    assert rows[0]["source_evse_id"] == "IBG-P-TYJE-B"
    assert rows[0]["operator_name"] == "In Balance grid, UAB"
    assert rows[0]["max_power_kw"] == 22.0
    assert rows[0]["tariff_ids"] == "418c326c-089e-11f0-a250-0a58a9feac0d"
    assert "availability_status" not in rows[0]


def test_lt_via_lietuva_ocpi_dynamic_parser_keeps_status_private_source():
    payload = {
        "timestamp": "2026-06-01T07:09:02Z",
        "data": [
            {
                "country_code": "LT",
                "party_id": "IBG",
                "id": 257,
                "coordinates": {"latitude": "54.7049540", "longitude": "25.2724750"},
                "evses": [
                    {
                        "uid": 12029,
                        "evse_id": "IBG-P-TYJE-B",
                        "status": "CHARGING",
                        "connectors": [{"id": 650305, "standard": "IEC_62196_T2", "max_electric_power": 22000}],
                        "last_updated": "2026-06-01T07:59:53",
                    },
                    {
                        "uid": 12030,
                        "evse_id": "IBG-P-TYJE-C",
                        "status": "REMOVED",
                        "connectors": [],
                    },
                ],
            }
        ],
    }

    rows = list(iter_lt_dynamic_rows_from_binary_stream(_json_stream(payload)))

    assert [row["source_uid"] for row in rows] == [LT_EV_STATUS_SOURCE_UID, LT_EV_STATUS_SOURCE_UID]
    assert [row["availability_status"] for row in rows] == ["occupied", "out_of_order"]
    assert rows[0]["source_observed_at"] == "2026-06-01T07:59:53"
    assert rows[1]["source_observed_at"] == "2026-06-01T07:09:02Z"


def test_lt_via_lietuva_datex_table_parser_uses_compound_refill_point_ids():
    payload = b"""
    <d2:payload xmlns:d2="https://datex2.eu/schema/3/d2Payload"
      xmlns:com="https://datex2.eu/schema/3/common"
      xmlns:egi="https://datex2.eu/schema/3/energyInfrastructure"
      xmlns:fac="https://datex2.eu/schema/3/facilities"
      xmlns:loc="https://datex2.eu/schema/3/locationReferencing">
      <com:publicationTime>2026-06-07T17:09:55+03:00</com:publicationTime>
      <egi:energyInfrastructureSite id="EGI-S-257" version="v1">
        <fac:name><com:values><com:value lang="lt">Z135 | Inbalance grid</com:value></com:values></fac:name>
        <fac:operator id="IBG">
          <fac:name><com:values><com:value lang="lt">In Balance grid, UAB</com:value></com:values></fac:name>
        </fac:operator>
        <fac:locationReference>
          <loc:pointByCoordinates>
            <loc:pointCoordinates>
              <loc:latitude>54.7049540</loc:latitude>
              <loc:longitude>25.2724750</loc:longitude>
            </loc:pointCoordinates>
          </loc:pointByCoordinates>
        </fac:locationReference>
        <egi:energyInfrastructureStation id="EGI-ST-257" version="v1">
          <egi:refillPoint id="0">
            <egi:connector>
              <egi:connectorType>iec62196T2</egi:connectorType>
              <egi:chargingMode>mode3AC1p</egi:chargingMode>
              <egi:maxPowerAtSocket>22000</egi:maxPowerAtSocket>
            </egi:connector>
          </egi:refillPoint>
        </egi:energyInfrastructureStation>
      </egi:energyInfrastructureSite>
    </d2:payload>
    """

    rows = list(iter_lt_rows_from_binary_stream(_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["source_uid"] == LT_EV_LOCATIONS_SOURCE_UID
    assert rows[0]["station_id"] == "lt:vialietuva:datex:egi-s-257"
    assert rows[0]["charger_id"] == "lt:vialietuva:datex:evse:egi-s-257-egi-st-257-0"
    assert rows[0]["source_evse_id"] == "EGI-S-257:EGI-ST-257:0"
    assert rows[0]["operator_name"] == "In Balance grid, UAB"
    assert rows[0]["max_power_kw"] == 22.0
    assert "availability_status" not in rows[0]


def test_lt_via_lietuva_datex_status_parser_uses_same_compound_ids():
    payload = b"""
    <d2:payload xmlns:d2="https://datex2.eu/schema/3/d2Payload"
      xmlns:com="https://datex2.eu/schema/3/common"
      xmlns:egi="https://datex2.eu/schema/3/energyInfrastructure"
      xmlns:fac="https://datex2.eu/schema/3/facilities">
      <com:publicationTime>2026-06-07T17:08:42+03:00</com:publicationTime>
      <egi:energyInfrastructureSiteStatus id="EGI-S-257" version="v1">
        <fac:lastUpdated>2026-06-07T17:08:39+03:00</fac:lastUpdated>
        <egi:energyInfrastructureStationStatus id="EGI-ST-257" version="v1">
          <egi:isAvailable>true</egi:isAvailable>
          <egi:refillPointStatus id="0" connectorIndex="1">
            <egi:status>CHARGING</egi:status>
          </egi:refillPointStatus>
          <egi:refillPointStatus id="1" connectorIndex="2">
            <egi:status>REMOVED</egi:status>
          </egi:refillPointStatus>
        </egi:energyInfrastructureStationStatus>
      </egi:energyInfrastructureSiteStatus>
    </d2:payload>
    """

    rows = list(iter_lt_dynamic_rows_from_binary_stream(_stream(payload)))

    assert [row["source_uid"] for row in rows] == [LT_EV_STATUS_SOURCE_UID, LT_EV_STATUS_SOURCE_UID]
    assert rows[0]["charger_id"] == "lt:vialietuva:datex:evse:egi-s-257-egi-st-257-0"
    assert rows[0]["availability_status"] == "occupied"
    assert rows[1]["availability_status"] == "out_of_order"
    assert rows[0]["source_observed_at"] == "2026-06-07T17:08:39+03:00"


def test_lt_public_source_specs_use_via_lietuva_datex_and_split_dynamic_status():
    assert PUBLIC_STATIC_SOURCES["lt"].url == "https://ev.vialietuva.lt/publicdata/EnergyInfrastructureTablePublication"
    assert PUBLIC_STATIC_SOURCES["lt"].task_kind == "parse_static_payload"
    assert PUBLIC_STATIC_SOURCES["lt"].content_type == "application/xml"
    assert PUBLIC_STATIC_SOURCES["lt-dynamic"].url == "https://ev.vialietuva.lt/publicdata/EnergyInfrastructureStatusPublication"
    assert PUBLIC_STATIC_SOURCES["lt-dynamic"].task_kind == "parse_dynamic_payload"
    assert PUBLIC_STATIC_SOURCES["lt-dynamic"].source_kind == "open_dynamic_no_auth_source"


def test_lt_via_lietuva_fetch_headers_accept_clearance_cookie(monkeypatch):
    monkeypatch.setenv("LT_VIA_LIETUVA_COOKIE", "cf_clearance=test-token")
    monkeypatch.setenv("LT_VIA_LIETUVA_USER_AGENT", "Mozilla/5.0 test")

    headers = fetch_eu._source_request_headers(PUBLIC_STATIC_SOURCES["lt"])

    assert headers["Cookie"] == "cf_clearance=test-token"
    assert headers["User-Agent"] == "Mozilla/5.0 test"
    assert headers["Referer"] == "https://ev.vialietuva.lt/en/data-provision"


def test_lt_via_lietuva_fetch_headers_wrap_bare_clearance_secret(monkeypatch, tmp_path):
    cookie_file = tmp_path / "lt_cookie.txt"
    cookie_file.write_text("bare-token", encoding="utf-8")
    monkeypatch.delenv("LT_VIA_LIETUVA_COOKIE", raising=False)
    monkeypatch.setenv("LT_VIA_LIETUVA_COOKIE_FILE", str(cookie_file))

    headers = fetch_eu._source_request_headers(PUBLIC_STATIC_SOURCES["lt-dynamic"])

    assert headers["Cookie"] == "cf_clearance=bare-token"
    assert headers["User-Agent"].startswith("Mozilla/5.0")


def test_lt_via_lietuva_cloudflare_challenge_is_explicit(monkeypatch):
    monkeypatch.delenv("LT_VIA_LIETUVA_COOKIE", raising=False)
    headers = Message()
    headers["cf-mitigated"] = "challenge"
    error = urllib.error.HTTPError(
        PUBLIC_STATIC_SOURCES["lt"].url,
        403,
        "Forbidden",
        headers,
        io.BytesIO(b"<!doctype html><title>Just a moment...</title>"),
    )

    with pytest.raises(fetch_eu.CloudflareChallengeError) as raised:
        fetch_eu._raise_cloudflare_challenge_if_present(
            source_key="lt",
            url=PUBLIC_STATIC_SOURCES["lt"].url,
            request_headers={},
            error=error,
        )

    message = str(raised.value)
    assert "cloudflare_challenge:lt:403:clearance_cookie_missing" in message
    assert "LT_VIA_LIETUVA_COOKIE_FILE" in message


def test_lt_via_lietuva_static_fetch_uses_tracked_fallback_after_cloudflare(monkeypatch, tmp_path, capsys):
    fallback_path = tmp_path / "lt-EnergyInfrastructureTablePublication.xml"
    fallback_payload = b"""
    <d2:payload xmlns:d2="https://datex2.eu/schema/3/d2Payload"
      xmlns:com="https://datex2.eu/schema/3/common"
      xmlns:egi="https://datex2.eu/schema/3/energyInfrastructure"
      xmlns:fac="https://datex2.eu/schema/3/facilities"
      xmlns:loc="https://datex2.eu/schema/3/locationReferencing">
      <com:publicationTime>2026-06-07T17:09:55+03:00</com:publicationTime>
      <egi:energyInfrastructureSite id="EGI-S-257" version="v1">
        <fac:locationReference>
          <loc:pointByCoordinates>
            <loc:pointCoordinates>
              <loc:latitude>54.7049540</loc:latitude>
              <loc:longitude>25.2724750</loc:longitude>
            </loc:pointCoordinates>
          </loc:pointByCoordinates>
        </fac:locationReference>
        <egi:energyInfrastructureStation id="EGI-ST-257" version="v1">
          <egi:refillPoint id="0">
            <egi:connector>
              <egi:connectorType>iec62196T2</egi:connectorType>
              <egi:maxPowerAtSocket>22000</egi:maxPowerAtSocket>
            </egi:connector>
          </egi:refillPoint>
        </egi:energyInfrastructureStation>
      </egi:energyInfrastructureSite>
    </d2:payload>
    """
    fallback_path.write_bytes(fallback_payload)
    monkeypatch.setenv(fetch_eu.LT_VIA_LIETUVA_STATIC_FALLBACK_ENV, str(fallback_path))

    def _raise_cloudflare(*_args, **_kwargs):
        raise fetch_eu.CloudflareChallengeError("cloudflare_challenge:lt:403")

    monkeypatch.setattr(fetch_eu, "_fetch_to_file", _raise_cloudflare)

    result = fetch_eu._fetch_source(
        config=AppConfig(raw_payload_dir=tmp_path / "raw"),
        spec=PUBLIC_STATIC_SOURCES["lt"],
        timeout_seconds=1,
    )

    assert result.payload_sha256 == hashlib.sha256(fallback_payload).hexdigest()
    assert result.content_type == "application/xml"
    assert result.final_url == fallback_path.resolve().as_uri()
    with result.payload_path.open("rb") as handle:
        rows = list(iter_lt_rows_from_binary_stream(handle))
    assert rows[0]["station_id"] == "lt:vialietuva:datex:egi-s-257"
    assert "using tracked static fallback" in capsys.readouterr().err


def test_lu_keeps_connector_count_as_one_station_placeholder_charger():
    payload = {
        "features": [
            {
                "id": "feature-1",
                "properties": {
                    "inspireid_identifier_localid": "RoadNode.1",
                    "gml_description": "Station A | 4 connectors with 22 kW and Type 2 connector",
                },
                "geometry": {"coordinates": [6.13, 49.61]},
            }
        ]
    }

    rows = list(iter_lu_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["charger_id"] == "lu:data-public:station-placeholder:roadnode-1"
    assert rows[0]["source_evse_id"] == "RoadNode.1|station-placeholder"
    assert rows[0]["connector_id"] == ""
    assert rows[0]["connector_types"] == "Type 2"


def test_mt_skips_features_without_valid_geometry():
    payload = {
        "features": [
            {
                "properties": {"Ref": "M-OK-001", "Name": "Valid", "State": "Available"},
                "geometry": {"coordinates": [14.48, 35.9]},
            },
            {"properties": {"Ref": "M-BAD-001"}, "geometry": None},
        ]
    }

    rows = list(iter_mt_rows_from_binary_stream(_json_stream(payload)))

    assert len(rows) == 1
    assert rows[0]["station_id"] == "mt:egis:m-ok-001"
    assert rows[0]["charger_id"] == "mt:egis:point:m-ok-001"
